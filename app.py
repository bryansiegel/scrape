from flask import Flask, render_template, request, jsonify, send_file, Response, stream_with_context
import mysql.connector
from mysql.connector import errorcode
import os
import re
import io
import base64
import subprocess
import openpyxl

# Load .env file if present (no external dependency needed)
_env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
if os.path.exists(_env_path):
    with open(_env_path) as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith('#') and '=' in _line:
                _k, _v = _line.split('=', 1)
                os.environ.setdefault(_k.strip(), _v.strip())
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# Single-scrape HTML cache — stores last scrape's page HTML keyed by URL
_html_cache = {}

# PDF-to-image cache — stores last conversion batch's rendered pages keyed by PDF URL.
# Each entry: {'source': str, 'pages': [png_bytes, ...], 'error': str|None, 'basename': str}
_pdf_image_cache = {}

# Accessibility Checker caches — reset at the start of each new scan/check batch.
_a11y_alt_findings = []   # [{'src','alt','hasAlt','source','pageUrl'}, ...]
_a11y_pdf_results  = {}   # pdf_url -> result dict from check_pdf_accessibility

# Tracking/analytics script detection patterns — Google + Facebook Pixel only.
TRACKER_PATTERNS = [
    ('Google Analytics 4',         ['gtag/js?id=G-', "gtag('config", "gtag('event"]),
    ('Google Universal Analytics', ['google-analytics.com/analytics.js', "ga('create"]),
    ('Google Tag Manager',         ['googletagmanager.com/gtm.js']),
    ('Facebook Pixel',             ['connect.facebook.net/en_US/fbevents.js', "fbq('init"]),
]

# Lazily-initialized spell checker singleton — loading the dictionary is too
# slow to redo on every page of an SEO crawl, so build it once and reuse it.
_spell_checker = None

def get_spell_checker():
    global _spell_checker
    if _spell_checker is None:
        from spellchecker import SpellChecker
        _spell_checker = SpellChecker()
    return _spell_checker


def get_ollama_url():
    return os.getenv('OLLAMA_URL', 'http://localhost:11434').rstrip('/')


def list_ollama_models(base_url):
    """Returns (running_model_names, installed_model_names) — best-effort, never raises."""
    import requests as req

    running = []
    try:
        r = req.get(f"{base_url}/api/ps", timeout=5)
        if r.ok:
            running = [m.get('name') for m in r.json().get('models', []) if m.get('name')]
    except Exception:
        pass

    installed = []
    try:
        r = req.get(f"{base_url}/api/tags", timeout=5)
        if r.ok:
            installed = [m.get('name') for m in r.json().get('models', []) if m.get('name')]
    except Exception:
        pass

    return running, installed


def url_to_filename(url, ext='.txt'):
    """Turn a page URL into a safe filename like 'subdomain--path--to--page.txt'."""
    from urllib.parse import urlparse

    parsed = urlparse(url)
    host = parsed.netloc.lower()
    if host.startswith('www.'):
        host = host[4:]
    if host.endswith('.ccsd.net'):
        host = host[:-len('.ccsd.net')]
    elif host == 'ccsd.net':
        host = 'ccsd'

    path = parsed.path.strip('/')
    name = f"{host}--{path.replace('/', '--')}" if path else host
    name = re.sub(r'[^a-zA-Z0-9\-_.]', '_', name)
    return f"{name}{ext}"


def pdf_url_to_basename(url):
    """Turn a PDF URL into a safe filename stem (no extension), e.g. for image output names."""
    from urllib.parse import urlparse, unquote

    name = unquote(urlparse(url).path.rsplit('/', 1)[-1]) or 'document'
    if name.lower().endswith('.pdf'):
        name = name[:-4]
    name = re.sub(r'[^a-zA-Z0-9\-_.]', '_', name)
    return name or 'document'


def stack_pages_into_one_image(page_pngs):
    """Stack a PDF's rendered pages top-to-bottom into a single tall PNG."""
    if len(page_pngs) == 1:
        return page_pngs[0]

    from PIL import Image

    pages  = [Image.open(io.BytesIO(png)).convert('RGB') for png in page_pngs]
    width  = max(p.width for p in pages)
    height = sum(p.height for p in pages)
    combined = Image.new('RGB', (width, height), 'white')

    y = 0
    for p in pages:
        combined.paste(p, ((width - p.width) // 2, y))
        y += p.height

    out = io.BytesIO()
    combined.save(out, format='PNG')
    return out.getvalue()


def resolve_img_src(img, base_url=None):
    """Get the real image URL for an <img> tag — a plain src/data-src/data-lazy-src,
    or (Finalsite CMS) the largest size baked into a data-image-sizes JSON blob."""
    from urllib.parse import urljoin

    src = (img.get('src') or img.get('data-src') or img.get('data-lazy-src') or '').strip()
    if src:
        return urljoin(base_url, src) if base_url else src

    raw_sizes = img.get('data-image-sizes')
    if raw_sizes:
        import json
        from urllib.parse import unquote
        try:
            sizes = json.loads(unquote(raw_sizes))
            best = max(sizes, key=lambda s: s.get('width', 0))
            if best.get('url'):
                return best['url']
        except Exception:
            pass

    return None


def extract_content_html(html_text, base_url=None):
    """Return only the main content HTML, stripping nav, header, footer, sidebars, and scripts."""
    from bs4 import BeautifulSoup, Tag, Comment

    try:
        soup = BeautifulSoup(html_text, 'html.parser')

        # Strip HTML comments outright — frameworks like Wix/React leave hydration
        # markers such as <!--$--> and <!--/$--> scattered through the DOM, which
        # otherwise survive as stray "$" / "/$" text once tags are unwrapped later.
        for c in soup.find_all(string=lambda t: isinstance(t, Comment)):
            c.extract()

        STRIP_TAGS = ['nav', 'header', 'footer', 'aside', 'script', 'style', 'noscript', 'iframe']
        STRIP_ATTRS = [
            'nav', 'navigation', 'navbar', 'menu', 'sidebar', 'side-bar',
            'site-header', 'site-footer', 'page-header', 'page-footer',
            'breadcrumb', 'pagination', 'widget', 'advertisement', 'banner',
            'cookie', 'popup', 'modal', 'social', 'share', 'related',
            'comment', 'toolbar', 'skip-link',
        ]
        CONTENT_SELECTORS = [
            'main', '[role="main"]', 'article',
            '#content', '#main-content', '#page-content', '#main', '#primary', '#body-content',
            '.content-wrap', '.content-area', '.main-content', '.page-content',
            '.entry-content', '.post-content', '.article-content', '.site-content',
            '.content', '.main',
        ]

        target = None
        for sel in CONTENT_SELECTORS:
            target = soup.select_one(sel)
            if target:
                break
        if target is None:
            target = soup.find('body') or soup

        # Collect then decompose — avoids accessing already-destroyed children
        # when iterating a pre-built list that includes their descendants.
        for tag in STRIP_TAGS:
            for elem in target.find_all(tag):
                elem.decompose()

        to_remove = []
        for elem in target.find_all(True):
            if not isinstance(elem, Tag):
                continue
            try:
                attrs = ' '.join(filter(None, [
                    elem.get('id') or '',
                    ' '.join(elem.get('class') or []),
                ])).lower()
                if any(p in attrs for p in STRIP_ATTRS):
                    to_remove.append(elem)
            except Exception:
                pass
        for elem in to_remove:
            try:
                elem.decompose()
            except Exception:
                pass

        # Strip class and style attributes from every remaining element
        for elem in target.find_all(True):
            if isinstance(elem, Tag):
                elem.attrs.pop('class', None)
                elem.attrs.pop('style', None)

        # Resolve each <img> to a real, absolute src (handling lazy-load
        # attributes and Finalsite's data-image-sizes CDN blobs) — drop any
        # image we can't resolve a URL for.
        for img in target.find_all('img'):
            resolved = resolve_img_src(img, base_url)
            if resolved:
                img['src'] = resolved
            else:
                img.decompose()

        return str(target)
    except Exception:
        # If anything goes wrong, fall back to returning the raw HTML
        return html_text


def clean_export_html(html_text):
    """Strip content HTML down to headings, paragraphs, links, lists, and tables only,
    then reformat it with each block tag on its own line and no stray whitespace."""
    from bs4 import BeautifulSoup, NavigableString, Comment

    ALLOWED_TAGS = {
        'h1', 'h2', 'h3', 'h4', 'h5', 'h6',
        'p', 'a', 'img',
        'ul', 'ol', 'li',
        'table', 'thead', 'tbody', 'tfoot', 'tr', 'td', 'th',
    }
    ALLOWED_ATTRS = {'a': {'href'}, 'img': {'src', 'alt'}}

    # Structural tags whose direct children are themselves tags (never meaningful
    # text) — each child gets its own output line. Everything else (headings,
    # paragraphs, list items, cells) renders its inline content on one line.
    CONTAINER_TAGS = {'ul', 'ol', 'table', 'thead', 'tbody', 'tfoot', 'tr'}
    LEAF_TAGS = {'h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'p', 'li', 'td', 'th'}

    # UI chrome that carries no page content of its own (form controls, decorative
    # media, icons) — delete these outright rather than unwrapping, or their
    # screen-reader-only label text (e.g. "pause slideshow") leaks into the output.
    # Note: <figure> is deliberately NOT here — it's just a wrapper and often
    # holds the <img> we want to keep, so it gets unwrapped instead, below.
    CHROME_TAGS = {
        'button', 'input', 'select', 'option', 'textarea', 'label', 'form',
        'svg', 'fieldset', 'legend', 'figcaption',
        'video', 'audio', 'canvas',
    }
    SR_ONLY_HINTS = ('sr-only', 'sronly', 'visually-hidden', 'visuallyhidden', 'screen-reader')

    try:
        soup = BeautifulSoup(html_text, 'html.parser')

        # Strip HTML comments outright — e.g. Wix/React hydration markers like
        # <!--$--> and <!--/$--> — before they can be kept as stray text below.
        for c in soup.find_all(string=lambda t: isinstance(t, Comment)):
            c.extract()

        for elem in soup.find_all(True):
            # A tag whose ancestor was already decomposed earlier in this same
            # loop (e.g. a <button><svg>...) is a dead object — .attrs is None
            # and .get() raises, which was silently swallowed by the outer
            # except and made this whole function a no-op on real pages.
            if elem.decomposed:
                continue
            if elem.name in CHROME_TAGS or elem.get('aria-hidden') == 'true':
                elem.decompose()
                continue
            attrs = ' '.join(filter(None, [
                elem.get('id') or '',
                ' '.join(elem.get('class') or []),
            ])).lower()
            if any(hint in attrs for hint in SR_ONLY_HINTS):
                elem.decompose()

        for elem in soup.find_all(True):
            if elem.name not in ALLOWED_TAGS:
                elem.unwrap()

        for elem in soup.find_all(True):
            keep = ALLOWED_ATTRS.get(elem.name, set())
            elem.attrs = {k: v for k, v in elem.attrs.items() if k in keep}

        # Drop links that lost their href, and images that lost their src,
        # and any tags left empty after unwrapping.
        for a in soup.find_all('a'):
            if not a.get('href'):
                a.unwrap()

        for img in soup.find_all('img'):
            if not img.get('src'):
                img.decompose()

        for tag_name in ('li', 'p', 'td', 'th'):
            for elem in soup.find_all(tag_name):
                if not elem.get_text(strip=True):
                    elem.decompose()

        # Collapse all internal whitespace runs (newlines, tabs, repeated spaces
        # from the original page's indentation) down to single spaces.
        for node in soup.find_all(string=True):
            collapsed = re.sub(r'\s+', ' ', str(node))
            node.replace_with(NavigableString(collapsed))

        # Container tags (ul/ol/table/thead/tbody/tfoot/tr) only ever hold other
        # tags — any text between their children is leftover indentation, so drop it.
        for elem in list(soup.find_all(CONTAINER_TAGS)) + [soup]:
            for child in list(elem.contents):
                if isinstance(child, NavigableString) and not child.strip():
                    child.extract()

        # Leaf/inline tags keep their text, but trim the whitespace that used to
        # separate them from sibling tags in the original markup.
        for elem in soup.find_all(list(LEAF_TAGS) + ['a']):
            if elem.contents and isinstance(elem.contents[0], NavigableString):
                elem.contents[0].replace_with(elem.contents[0].lstrip())
            if elem.contents and isinstance(elem.contents[-1], NavigableString):
                elem.contents[-1].replace_with(elem.contents[-1].rstrip())

        # Serialize with each container/leaf tag on its own line and no added
        # indentation — inline tags (links) stay inline within their parent line.
        def render(node):
            lines = []
            for child in node.contents:
                if isinstance(child, NavigableString):
                    if child.strip():
                        lines.append(str(child))
                    continue
                if child.name in CONTAINER_TAGS:
                    lines.append(f'<{child.name}>')
                    lines.extend(render(child))
                    lines.append(f'</{child.name}>')
                else:
                    lines.append(str(child))
            return lines

        return '\n'.join(render(soup))
    except Exception:
        return html_text


def get_keyword_trends(phrases):
    """Return {phrase: avg_interest_0_to_100_or_None} via pytrends (best-effort, never raises)."""
    if not phrases:
        return {}
    try:
        from pytrends.request import TrendReq
        import time
        scores = {}
        batches = [phrases[i:i + 5] for i in range(0, len(phrases), 5)]
        for idx, batch in enumerate(batches):
            try:
                pt = TrendReq(hl='en-US', tz=360, timeout=(10, 30), retries=1, backoff_factor=0.5)
                pt.build_payload(batch, timeframe='today 12-m', geo='US')
                df = pt.interest_over_time()
                for phrase in batch:
                    if df is not None and not df.empty and phrase in df.columns:
                        scores[phrase] = int(round(float(df[phrase].mean())))
                    else:
                        scores[phrase] = None
            except Exception:
                for phrase in batch:
                    scores[phrase] = None
            if idx < len(batches) - 1:
                time.sleep(2.5)
        return scores
    except Exception:
        return {p: None for p in phrases}


# ---------------------------------------------------------------------------
# Accessibility Checker helpers
# ---------------------------------------------------------------------------

_axe_script_cache = None

def get_axe_script():
    """Read the vendored axe-core build once and cache it in memory."""
    global _axe_script_cache
    if _axe_script_cache is None:
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'vendor', 'axe.min.js')
        with open(path, 'r', encoding='utf-8') as f:
            _axe_script_cache = f.read()
    return _axe_script_cache


_axe_playwright = None
_axe_browser = None

def get_axe_browser():
    """Lazily launch a single shared headless Chromium instance, reused for an entire scan
    rather than relaunched per page (launching is the expensive part)."""
    global _axe_playwright, _axe_browser
    if _axe_browser is None:
        from playwright.sync_api import sync_playwright
        _axe_playwright = sync_playwright().start()
        _axe_browser = _axe_playwright.chromium.launch()
    return _axe_browser


def close_axe_browser():
    global _axe_playwright, _axe_browser
    if _axe_browser is not None:
        try:
            _axe_browser.close()
        except Exception:
            pass
        _axe_browser = None
    if _axe_playwright is not None:
        try:
            _axe_playwright.stop()
        except Exception:
            pass
        _axe_playwright = None


AXE_IMPACT_PENALTY = {'critical': 25, 'serious': 15, 'moderate': 7, 'minor': 3}

def run_axe_on_url(url, timeout_ms=20000):
    """Render the page in a headless browser and run axe-core against it — the same
    WCAG rule engine class WAVE is built on. Returns
    {'violations': [...], 'passCount': int, 'error': str|None} — best-effort, never raises."""
    page = None
    try:
        browser = get_axe_browser()
        page = browser.new_page()
        page.goto(url, timeout=timeout_ms, wait_until='load')
        page.add_script_tag(content=get_axe_script())
        results = page.evaluate(
            "() => axe.run(document, {runOnly: ['wcag2a', 'wcag2aa', 'wcag21aa']})"
        )
        violations = [
            {
                'id': v.get('id'),
                'impact': v.get('impact') or 'minor',
                'help': v.get('help'),
                'helpUrl': v.get('helpUrl'),
                'nodeCount': len(v.get('nodes', [])),
            }
            for v in results.get('violations', [])
        ]
        return {'violations': violations, 'passCount': len(results.get('passes', [])), 'error': None}
    except Exception as e:
        return {'violations': [], 'passCount': 0, 'error': str(e)}
    finally:
        if page is not None:
            try:
                page.close()
            except Exception:
                pass


def score_from_axe(axe_result):
    """0-100: 100 minus a per-violation-node penalty weighted by impact severity."""
    if axe_result.get('error'):
        return None
    score = 100
    for v in axe_result.get('violations', []):
        penalty = AXE_IMPACT_PENALTY.get(v.get('impact'), 3)
        score -= penalty * max(1, v.get('nodeCount', 1))
    return max(0, score)


def extract_alt_tag_findings(soup, page_url):
    """Return [{'src','alt','hasAlt','source','pageUrl'}, ...] for every <img> on the page,
    'source' being the image tag's own raw markup so a missing/bad alt can be located in code."""
    findings = []
    for img in soup.find_all('img'):
        full_img = resolve_img_src(img, page_url)
        if not full_img or not full_img.startswith('http'):
            continue
        alt = img.get('alt')
        has_alt = alt is not None and alt.strip() != ''
        source_html = get_snippet(img)
        findings.append({
            'src': full_img, 'alt': alt or '', 'hasAlt': has_alt,
            'source': source_html, 'pageUrl': page_url,
        })
    return findings


def get_snippet(element, limit=600):
    try:
        html = str(element)
        return html[:limit] + ('…' if len(html) > limit else '')
    except Exception:
        return ''


PDF_HEADING_TAGS = {'/H', '/H1', '/H2', '/H3', '/H4', '/H5', '/H6'}

def check_pdf_accessibility(pdf_bytes, pdf_url, use_ai=False, model=None, ollama_url=None, timeout=120):
    """Machine-verifiable PDF/UA structural checks via pikepdf (tagged, language, title,
    figure alt text, heading structure) — the same category of rules check.axes4.com applies.
    Never raises; failures are reported in result['error']."""
    import pikepdf

    result = {
        'url': pdf_url, 'tagged': False, 'hasLang': False, 'hasTitle': False,
        'figureCount': 0, 'figuresWithAlt': 0, 'hasHeadings': False,
        'score': 0, 'aiNotes': None, 'error': None,
    }
    try:
        with pikepdf.open(io.BytesIO(pdf_bytes)) as pdf:
            root = pdf.Root

            mark_info = root.get('/MarkInfo')
            result['tagged'] = bool(mark_info is not None and bool(mark_info.get('/Marked', False)))

            lang = root.get('/Lang')
            result['hasLang'] = bool(lang is not None and str(lang).strip())

            title = None
            try:
                if pdf.docinfo is not None and '/Title' in pdf.docinfo:
                    title = str(pdf.docinfo['/Title'])
            except Exception:
                title = None
            if not title:
                try:
                    with pdf.open_metadata() as meta:
                        title = meta.get('dc:title')
                except Exception:
                    title = None
            result['hasTitle'] = bool(title and str(title).strip())

            alt_texts = []

            def walk(elem, depth=0):
                if depth > 60 or elem is None:
                    return
                try:
                    s_type = elem.get('/S')
                    tag = f'/{s_type}' if s_type is not None else None
                except Exception:
                    tag = None

                if tag == '/Figure':
                    result['figureCount'] += 1
                    try:
                        alt = elem.get('/Alt')
                    except Exception:
                        alt = None
                    if alt and str(alt).strip():
                        result['figuresWithAlt'] += 1
                        alt_texts.append(str(alt).strip())
                elif tag in PDF_HEADING_TAGS:
                    result['hasHeadings'] = True

                try:
                    kids = elem.get('/K')
                except Exception:
                    kids = None
                if kids is None:
                    return
                try:
                    if isinstance(kids, pikepdf.Array):
                        for k in kids:
                            if isinstance(k, pikepdf.Dictionary):
                                walk(k, depth + 1)
                    elif isinstance(kids, pikepdf.Dictionary):
                        walk(kids, depth + 1)
                except Exception:
                    pass

            struct_root = root.get('/StructTreeRoot')
            if struct_root is not None:
                walk(struct_root)

            score = 0
            score += 30 if result['tagged'] else 0
            score += 15 if result['hasLang'] else 0
            score += 10 if result['hasTitle'] else 0
            if result['figureCount'] > 0:
                score += round(25 * (result['figuresWithAlt'] / result['figureCount']))
            elif result['tagged']:
                score += 25  # tagged with no figures to caption — nothing to penalize
            score += 20 if result['hasHeadings'] else 0
            result['score'] = min(100, score)

            if use_ai and alt_texts:
                result['aiNotes'] = get_pdf_alt_quality_notes(alt_texts, model, ollama_url, timeout)

    except Exception as e:
        result['error'] = str(e)

    return result


def get_pdf_alt_quality_notes(alt_texts, model, ollama_url, timeout):
    """Ask local Ollama to flag non-descriptive alt text among a PDF's figures — best-effort,
    same request pattern as the SEO checker's Ollama calls."""
    import requests as req
    if not model:
        return None
    sample = alt_texts[:20]
    prompt = (
        "You are an accessibility reviewer. Below is a list of alt-text descriptions "
        "attached to figures in a PDF document. Identify any that are NOT descriptive "
        "(e.g. a filename, 'image', 'picture1.jpg', or otherwise meaningless to a screen "
        "reader user). Respond with one short line per non-descriptive entry in the form "
        "'BAD: <text>' — if all entries are descriptive, respond with exactly 'OK'.\n\n"
        + "\n".join(f"- {t}" for t in sample)
    )
    try:
        r = req.post(
            f"{ollama_url}/api/generate",
            json={'model': model, 'prompt': prompt, 'stream': False},
            timeout=timeout,
        )
        r.raise_for_status()
        return r.json().get('response', '').strip()
    except Exception:
        return None


def compute_overall_a11y_score(page_scores, pdf_scores):
    scores = [s for s in page_scores if s is not None] + [s for s in pdf_scores if s is not None]
    if not scores:
        return None
    return round(sum(scores) / len(scores))


BAND_COLORS = {'good': '#198754', 'mid': '#ffc107', 'bad': '#dc3545', 'na': '#adb5bd'}

def score_band(score):
    if score is None:
        return 'na'
    if score >= 80:
        return 'good'
    if score >= 50:
        return 'mid'
    return 'bad'

def pct_band_inverse(pct):
    """For a 'percentage that is bad' metric (e.g. % of PDFs inaccessible) — low is good."""
    if pct is None:
        return 'na'
    if pct <= 20:
        return 'good'
    if pct <= 50:
        return 'mid'
    return 'bad'


# Database configuration — values loaded from .env
DB_CONFIG = {
    'host':     os.getenv('DB_HOST', '127.0.0.1'),
    'user':     os.getenv('DB_USER', 'root'),
    'password': os.getenv('DB_PASSWORD', ''),
    'database': os.getenv('DB_NAME', 'scrape'),
}

def get_db_connection():
    try:
        connection = mysql.connector.connect(**DB_CONFIG)
        return connection
    except mysql.connector.Error as err:
        print(f"Database connection error: {err}")
        return None

def search_txt_files(search_term, category=None):
    """Search through .txt files for matching URLs"""
    results = []
    
    # Define file mappings
    file_mappings = {
        'departments': 'scraper_departments.txt',
        'divisions': 'scraper_divisions.txt',
        'general': 'scraped_links.txt',
        'drive': 'scraper_drive_links.txt',
        'googleSites': 'scraped_google_sites.txt',
        'pdf': 'scraped_pdf.txt'
    }
    
    # Determine which files to search
    files_to_search = {}
    if category and category != 'all' and category in file_mappings:
        files_to_search[category] = file_mappings[category]
    else:
        files_to_search = file_mappings
    
    for cat, filename in files_to_search.items():
        filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
        if os.path.exists(filepath):
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    for line_num, line in enumerate(f, 1):
                        raw = line.strip()
                        if not raw:
                            continue
                        if cat == 'pdf' and '|' in raw:
                            url, source_page = raw.split('|', 1)
                        else:
                            url = raw
                            source_page = None
                        if not search_term or search_term.lower() in url.lower():
                            result = {
                                'category': cat,
                                'url': url,
                                'source': 'txt_file',
                                'file': filename,
                                'line': line_num
                            }
                            if source_page:
                                result['source_page'] = source_page
                            results.append(result)
            except Exception as e:
                print(f"Error reading {filepath}: {e}")
    
    return results

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/data')
def get_data():
    category = request.args.get('category', 'all')
    search = request.args.get('search', '')
    
    all_results = []
    
    # Search database (pdf has no DB column — handled via txt file only)
    connection = get_db_connection() if category != 'pdf' else None
    if connection:
        cursor = connection.cursor()
        try:
            if category == 'all':
                query = """
                SELECT 'departments' as category, departments as url, 'database' as source 
                FROM pages WHERE departments IS NOT NULL AND departments != ''
                UNION ALL
                SELECT 'drive' as category, drive as url, 'database' as source 
                FROM pages WHERE drive IS NOT NULL AND drive != ''
                UNION ALL  
                SELECT 'divisions' as category, divisions as url, 'database' as source 
                FROM pages WHERE divisions IS NOT NULL AND divisions != ''
                UNION ALL
                SELECT 'general' as category, general as url, 'database' as source 
                FROM pages WHERE general IS NOT NULL AND general != ''
                UNION ALL
                SELECT 'googleSites' as category, googleSites as url, 'database' as source 
                FROM pages WHERE googleSites IS NOT NULL AND googleSites != ''
                """
            else:
                query = f"SELECT '{category}' as category, {category} as url, 'database' as source FROM pages WHERE {category} IS NOT NULL AND {category} != ''"
            
            if search:
                if category == 'all':
                    query += f" HAVING url LIKE %s"
                else:
                    query += f" AND {category} LIKE %s"
                
                cursor.execute(query, (f'%{search}%',))
            else:
                cursor.execute(query)
            
            db_results = cursor.fetchall()
            all_results.extend([{'category': row[0], 'url': row[1], 'source': row[2]} for row in db_results])
            
        except mysql.connector.Error as err:
            print(f"Database query error: {err}")
        finally:
            cursor.close()
            connection.close()
    
    # Search txt files
    txt_results = search_txt_files(search, category)
    all_results.extend(txt_results)
    
    # Remove duplicates while preserving order (database results first)
    seen_urls = set()
    unique_results = []
    for result in all_results:
        url = result['url']
        if url not in seen_urls:
            seen_urls.add(url)
            unique_results.append(result)
    
    return jsonify({'data': unique_results, 'count': len(unique_results)})

@app.route('/api/stats')
def get_stats():
    stats = {'departments': 0, 'drive': 0, 'divisions': 0, 'general': 0, 'googleSites': 0, 'pdf': 0, 'total': 0}
    
    # Count from database
    connection = get_db_connection()
    if connection:
        cursor = connection.cursor()
        try:
            categories = ['departments', 'drive', 'divisions', 'general', 'googleSites']
            # Note: 'pdf' has no DB column — counted from txt file only
            
            for category in categories:
                cursor.execute(f"SELECT COUNT(*) FROM pages WHERE {category} IS NOT NULL AND {category} != ''")
                count = cursor.fetchone()[0]
                stats[category] += count
        except mysql.connector.Error as err:
            print(f"Database stats error: {err}")
        finally:
            cursor.close()
            connection.close()
    
    # Count from txt files
    file_mappings = {
        'departments': 'scraper_departments.txt',
        'divisions': 'scraper_divisions.txt',
        'general': 'scraped_links.txt',
        'drive': 'scraper_drive_links.txt',
        'googleSites': 'scraped_google_sites.txt',
        'pdf': 'scraped_pdf.txt'
    }

    for category, filename in file_mappings.items():
        filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
        if os.path.exists(filepath):
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    file_count = sum(1 for line in f if line.strip())
                stats[category] += file_count
            except Exception as e:
                print(f"Error counting {filepath}: {e}")
    
    # Calculate total unique URLs
    all_urls = set()
    
    # Add database URLs
    connection = get_db_connection()
    if connection:
        cursor = connection.cursor()
        try:
            cursor.execute("""
                SELECT departments FROM pages WHERE departments IS NOT NULL AND departments != ''
                UNION ALL
                SELECT drive FROM pages WHERE drive IS NOT NULL AND drive != ''
                UNION ALL
                SELECT divisions FROM pages WHERE divisions IS NOT NULL AND divisions != ''
                UNION ALL
                SELECT general FROM pages WHERE general IS NOT NULL AND general != ''
                UNION ALL
                SELECT googleSites FROM pages WHERE googleSites IS NOT NULL AND googleSites != ''
            """)
            for (url,) in cursor.fetchall():
                if url:
                    all_urls.add(url.strip())
        except mysql.connector.Error as err:
            print(f"Database total count error: {err}")
        finally:
            cursor.close()
            connection.close()
    
    # Add txt file URLs (file_mappings already includes pdf)
    for cat, filename in file_mappings.items():
        filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
        if os.path.exists(filepath):
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    for line in f:
                        raw = line.strip()
                        if raw:
                            url = raw.split('|', 1)[0] if cat == 'pdf' else raw
                            all_urls.add(url)
            except Exception as e:
                print(f"Error reading {filepath} for total: {e}")
    
    stats['total'] = len(all_urls)
    return jsonify(stats)

@app.route('/api/autocomplete')
def get_autocomplete():
    query = request.args.get('q', '').lower()
    suggestions = set()
    
    # Get suggestions from txt files
    file_mappings = {
        'departments': 'scraper_departments.txt',
        'divisions': 'scraper_divisions.txt',
        'general': 'scraped_links.txt',
        'drive': 'scraper_drive_links.txt',
        'googleSites': 'scraped_google_sites.txt',
        'pdf': 'scraped_pdf.txt'
    }
    
    for filename in file_mappings.values():
        filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
        if os.path.exists(filepath):
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    for line in f:
                        url = line.strip()
                        if url:
                            # Extract meaningful parts for autocomplete
                            parts = re.split(r'[/\-_.]', url)
                            for part in parts:
                                if len(part) > 2 and not part.isdigit():
                                    suggestions.add(part.lower())
            except Exception as e:
                print(f"Error reading {filepath}: {e}")
    
    # Filter and return suggestions
    all_suggestions = sorted(list(suggestions))
    if query:
        filtered = [s for s in all_suggestions if query in s]
        return jsonify(filtered[:10])
    else:
        return jsonify(all_suggestions[:10])

@app.route('/api/export/pdf')
def export_pdf_excel():
    search = request.args.get('search', '')

    # Load all PDF entries from scraped_pdf.txt
    filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'scraped_pdf.txt')
    groups = {}  # source_page -> [pdf_url, ...]
    if os.path.exists(filepath):
        with open(filepath, 'r', encoding='utf-8') as f:
            for line in f:
                raw = line.strip()
                if not raw:
                    continue
                if '|' in raw:
                    pdf_url, source_page = raw.split('|', 1)
                else:
                    pdf_url, source_page = raw, '(unknown page)'
                if search and search.lower() not in pdf_url.lower() and search.lower() not in source_page.lower():
                    continue
                if source_page not in groups:
                    groups[source_page] = []
                groups[source_page].append(pdf_url)

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PDFs by Page'

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(fill_type='solid', fgColor='1F3864')
    page_font = Font(bold=True, size=10)
    page_fill = PatternFill(fill_type='solid', fgColor='D9E1F2')
    link_font = Font(color='1155CC', size=10)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 80

    ws.append(['#', 'Source Page', 'PDF URL'])
    for col in range(1, 4):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 20

    row_num = 2
    entry_num = 1
    for source_page, pdfs in sorted(groups.items()):
        # Source page group header
        if len(ws.merged_cells.ranges):
            pass
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
        cell = ws.cell(row=row_num, column=1)
        cell.value = f'Page: {source_page}'
        cell.font = page_font
        cell.fill = page_fill
        cell.alignment = left
        cell.border = border
        # Apply border to merged cells individually
        for col in range(2, 4):
            ws.cell(row=row_num, column=col).border = border
        ws.row_dimensions[row_num].height = 18
        row_num += 1

        for pdf_url in pdfs:
            ws.cell(row=row_num, column=1, value=entry_num).alignment = center
            ws.cell(row=row_num, column=1).border = border
            ws.cell(row=row_num, column=2, value=source_page).font = Font(size=10)
            ws.cell(row=row_num, column=2).alignment = left
            ws.cell(row=row_num, column=2).border = border
            pdf_cell = ws.cell(row=row_num, column=3, value=pdf_url)
            pdf_cell.font = link_font
            pdf_cell.alignment = left
            pdf_cell.border = border
            ws.row_dimensions[row_num].height = 16
            row_num += 1
            entry_num += 1

    # Freeze header row
    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='ccsd_pdfs.xlsx'
    )


@app.route('/api/export/scrape-pdfs', methods=['POST'])
def export_scrape_pdfs():
    data          = request.get_json(force=True)
    dl_name       = data.get('download_name', 'scrape_pdfs.xlsx')
    wide_headers  = data.get('wide_headers')   # present for SEO wide-pivot export
    wide_rows     = data.get('wide_rows')

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(fill_type='solid', fgColor='1F3864')
    link_font   = Font(color='1155CC', size=10)
    easy_font   = Font(color='1E7E34', size=10, bold=True)
    hard_font   = Font(color='B8430A', size=10, bold=True)
    center      = Alignment(horizontal='center', vertical='center')
    left        = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin        = Side(style='thin', color='BFBFBF')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PDFs'

    if wide_headers and wide_rows:
        # ── Wide pivot format (SEO checker): one row per source URL ──────────
        # Column layout: Source URL (col 1), then groups of [PDF URL, Pages, Difficulty]
        num_cols = len(wide_headers)
        ws.column_dimensions['A'].width = 70  # Source URL
        for i in range(1, num_cols):
            col_letter = chr(ord('A') + i)
            header_lc  = wide_headers[i].lower()
            if 'url' in header_lc:
                ws.column_dimensions[col_letter].width = 70
            elif 'pages' in header_lc:
                ws.column_dimensions[col_letter].width = 10
            else:  # Difficulty
                ws.column_dimensions[col_letter].width = 14

        # Header row
        for col_idx, hdr in enumerate(wide_headers, 1):
            cell           = ws.cell(row=1, column=col_idx, value=hdr)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            cell.border    = border
        ws.row_dimensions[1].height = 20

        for row_idx, row_data in enumerate(wide_rows, 2):
            for col_idx, val in enumerate(row_data, 1):
                cell        = ws.cell(row=row_idx, column=col_idx, value=val if val != '' else None)
                cell.border = border
                header_lc   = wide_headers[col_idx - 1].lower()
                if col_idx == 1:                        # Source URL
                    cell.font      = Font(size=10)
                    cell.alignment = left
                elif 'url' in header_lc:               # PDF URL columns
                    cell.font      = link_font
                    cell.alignment = left
                elif 'difficulty' in header_lc:         # Difficulty columns
                    vl = str(val).lower() if val else ''
                    cell.font      = easy_font if vl == 'easy' else (hard_font if vl == 'hard' else Font(size=10))
                    cell.alignment = center
                else:                                   # Pages columns
                    cell.font      = Font(size=10)
                    cell.alignment = center
            ws.row_dimensions[row_idx].height = 16

    else:
        # ── Original tall format (single-scrape) ─────────────────────────────
        pdfs      = data.get('pdfs', [])
        col_label = data.get('col_label', 'Source / File Name')
        extra_cols = data.get('extra_cols', [])

        col_widths = [6, 60, 80] + [16] * len(extra_cols)
        for i, w in enumerate(col_widths):
            ws.column_dimensions[chr(ord('A') + i)].width = w

        headers = ['#', col_label, 'PDF URL'] + extra_cols
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell           = ws.cell(row=1, column=col)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            cell.border    = border
        ws.row_dimensions[1].height = 20

        for i, pdf in enumerate(pdfs, 1):
            url   = pdf.get('url', '')
            label = pdf.get('label', '')

            ws.cell(row=i + 1, column=1, value=i).alignment = center
            ws.cell(row=i + 1, column=1).border             = border
            ws.cell(row=i + 1, column=1).font               = Font(size=10)

            c2 = ws.cell(row=i + 1, column=2, value=label)
            c2.font = Font(size=10); c2.alignment = left; c2.border = border

            c3 = ws.cell(row=i + 1, column=3, value=url)
            c3.font = link_font; c3.alignment = left; c3.border = border

            for j, col_name in enumerate(extra_cols):
                val  = pdf.get(col_name.lower(), '') or ''
                cell = ws.cell(row=i + 1, column=4 + j, value=val)
                cell.alignment = center
                cell.border    = border
                if col_name.lower() == 'difficulty':
                    vl = str(val).lower()
                    cell.font = easy_font if vl == 'easy' else (hard_font if vl == 'hard' else Font(size=10))
                else:
                    cell.font = Font(size=10)
            ws.row_dimensions[i + 1].height = 16

    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=dl_name,
    )


SCRAPER_SCRIPTS = {
    'all': 'all-scrape.py',
    'main': 'scraper.py',
    'html': 'scraper-html.py',
    'pdf': 'scraper_pdf.py',
    'drive': 'scraper_drive_links.py',
    'sites': 'scraper_google_sites.py',
    'database': 'add_to_database.py',
}

@app.route('/api/scrape/<scraper_type>')
def run_scraper(scraper_type):
    if scraper_type not in SCRAPER_SCRIPTS:
        def err():
            yield f"data: ERROR: Unknown scraper '{scraper_type}'\n\n"
            yield "data: __FAILURE__\n\n"
        return Response(stream_with_context(err()), mimetype='text/event-stream')

    script = SCRAPER_SCRIPTS[scraper_type]
    script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), script)

    if not os.path.exists(script_path):
        def err():
            yield f"data: ERROR: Script not found: {script}\n\n"
            yield "data: __FAILURE__\n\n"
        return Response(stream_with_context(err()), mimetype='text/event-stream')

    def generate():
        process = subprocess.Popen(
            ['python', '-u', script_path],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1
        )
        for line in process.stdout:
            yield f"data: {line.rstrip()}\n\n"
        process.wait()
        if process.returncode == 0:
            yield "data: __SUCCESS__\n\n"
        else:
            yield f"data: __FAILURE__ (exit code {process.returncode})\n\n"

    return Response(stream_with_context(generate()), mimetype='text/event-stream')


@app.route('/single-scrape')
def single_scrape_page():
    return render_template('single_scrape.html')


@app.route('/page-scrape')
def page_scrape_page():
    return render_template('page_scrape.html')


@app.route('/broken-link-checker')
def broken_link_checker_page():
    return render_template('broken_link_checker.html')


@app.route('/seo-checker')
def seo_checker_page():
    return render_template('seo_checker.html')


@app.route('/api/single-scrape')
def run_single_scrape():
    import requests as req
    from bs4 import BeautifulSoup
    from urllib.parse import urlparse, urljoin
    from collections import deque
    import time
    import random

    url           = request.args.get('url', '').strip()
    find_pdf      = request.args.get('pdf', 'true').lower() == 'true'
    find_sites    = request.args.get('sites', 'true').lower() == 'true'
    find_images   = request.args.get('images', 'true').lower() == 'true'
    find_tracking = request.args.get('tracking', 'true').lower() == 'true'
    find_external = request.args.get('external', 'true').lower() == 'true'

    def err(msg):
        yield f"data: {msg}\n\n"
        yield "data: __FAILURE__\n\n"

    if not url or not url.startswith('http'):
        return Response(stream_with_context(err('ERROR: Invalid or missing URL')), mimetype='text/event-stream')

    USER_AGENTS = [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15',
        'Mozilla/5.0 (X11; Linux x86_64; rv:120.0) Gecko/20100101 Firefox/120.0',
    ]

    def generate():
        global _html_cache
        _html_cache = {}  # clear previous scrape

        parsed_start = urlparse(url)
        base_domain  = parsed_start.netloc.lower().removeprefix('www.')

        visited        = set()
        queued         = {url}
        queue          = deque([url])
        found_pdfs     = set()
        found_sites    = set()
        found_images   = set()
        found_trackers = set()  # "tracker_name|code" keys — unique script code, site-wide
        found_externals = set()
        pages_crawled  = 0

        yield f"data: Starting crawl of {url}\n\n"
        yield f"data: Domain: {base_domain}\n\n"

        while queue:
            current = queue.popleft()
            if current in visited:
                continue
            visited.add(current)
            pages_crawled += 1

            yield f"data: PAGE: {pages_crawled}\n\n"
            yield f"data: [{pages_crawled}] {current}\n\n"

            try:
                headers = {
                    'User-Agent': random.choice(USER_AGENTS),
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
                    'Accept-Language': 'en-US,en;q=0.9',
                    'Connection': 'keep-alive',
                    'Upgrade-Insecure-Requests': '1',
                }
                resp = req.get(current, headers=headers, timeout=10, allow_redirects=True)
                if resp.status_code != 200:
                    yield f"data: Skipped (HTTP {resp.status_code}): {current}\n\n"
                    continue

                content_type = resp.headers.get('Content-Type', '')
                if 'html' not in content_type:
                    continue

                raw_html = resp.text
                _html_cache[current] = extract_content_html(raw_html, current)
                soup = BeautifulSoup(raw_html, 'html.parser')

                # Emit LINK with page title so the All Links tab can label it
                title_tag  = soup.find('title')
                page_title = title_tag.get_text().strip()[:100] if title_tag else ''
                yield f"data: LINK: {current}|{page_title}\n\n"

                # Walk all <a> tags
                for a in soup.find_all('a', href=True):
                    href = a['href'].strip()
                    if not href or href.startswith('mailto:') or href.startswith('javascript:'):
                        continue
                    full = urljoin(current, href).split('#')[0].rstrip('/')
                    if not full.startswith('http'):
                        continue

                    data_file = (a.get('data-file-name') or '').strip()
                    url_path  = full.lower().split('?')[0]
                    is_pdf    = url_path.endswith('.pdf') or data_file.lower().endswith('.pdf')
                    if find_pdf and is_pdf and full not in found_pdfs:
                        found_pdfs.add(full)
                        yield f"data: PDF: {full}|{current}\n\n"

                    if find_sites and 'sites.google.com' in full and full not in found_sites:
                        found_sites.add(full)
                        yield f"data: SITE: {full}|{current}\n\n"

                    link_domain = urlparse(full).netloc.lower().removeprefix('www.')
                    if link_domain == base_domain:
                        if full not in visited and full not in queued:
                            queued.add(full)
                            queue.append(full)
                    elif find_external and full not in found_externals:
                        found_externals.add(full)
                        rel = a.get('rel') or []
                        if isinstance(rel, str):
                            rel = rel.split()
                        nofollow = 'nofollow' in [r.lower() for r in rel]
                        is_ccsd = link_domain == 'ccsd.net' or link_domain.endswith('.ccsd.net')
                        yield f"data: EXTERNAL: {full}|{current}|{1 if nofollow else 0}|{1 if is_ccsd else 0}\n\n"

                # Collect images from <img> tags only (no external CDNs/other
                # sites) — except a page-embedded resource CDN (e.g. Finalsite's
                # data-image-sizes-hosted photos) still counts as first-party
                # since it's this site's own asset host, not a foreign reference.
                if find_images:
                    for img in soup.find_all('img'):
                        plain_src = (img.get('src') or img.get('data-src') or
                                     img.get('data-lazy-src') or '').strip()
                        is_first_party_cdn = not plain_src and img.get('data-image-sizes')
                        full_img = resolve_img_src(img, current)
                        if not full_img or not full_img.startswith('http') or full_img in found_images:
                            continue
                        if not is_first_party_cdn:
                            img_domain = urlparse(full_img).netloc.lower().removeprefix('www.')
                            if img_domain != base_domain:
                                continue
                        found_images.add(full_img)
                        yield f"data: IMG: {full_img}|{current}\n\n"

                # Detect tracking/analytics scripts — capture each script's actual
                # code (src URL for external scripts, full body for inline ones)
                # and dedupe site-wide so an identical snippet on every page only
                # shows up once.
                if find_tracking:
                    for script in soup.find_all('script'):
                        src  = (script.get('src') or '').strip()
                        text = (script.get_text() or '').strip()
                        haystack = f"{src} {text}".lower()
                        code = src or text
                        if not code:
                            continue
                        for name, patterns in TRACKER_PATTERNS:
                            if any(p.lower() in haystack for p in patterns):
                                key = f"{name}|{code}"
                                if key not in found_trackers:
                                    found_trackers.add(key)
                                    kind = 'src' if src else 'inline'
                                    code_b64 = base64.b64encode(code.encode('utf-8')).decode('ascii')
                                    yield f"data: TRACKER: {name}|{current}|{kind}|{code_b64}\n\n"
                                break

            except Exception as e:
                yield f"data: Error ({current}): {e}\n\n"

            time.sleep(0.15)

        yield f"data: \n\n"
        yield (
            f"data: Done. {pages_crawled} page(s) crawled. "
            f"{len(found_pdfs)} PDF(s), {len(found_sites)} Google Sites, "
            f"{len(found_images)} image(s), {len(found_trackers)} tracker(s), "
            f"{len(found_externals)} external link(s) found.\n\n"
        )
        yield "data: __SUCCESS__\n\n"

    return Response(stream_with_context(generate()), mimetype='text/event-stream')


@app.route('/api/page-scrape')
def run_page_scrape():
    import requests as req
    from bs4 import BeautifulSoup
    from urllib.parse import urljoin
    import random

    url           = request.args.get('url', '').strip()
    find_pdf      = request.args.get('pdf', 'true').lower() == 'true'
    find_sites    = request.args.get('sites', 'true').lower() == 'true'
    find_images   = request.args.get('images', 'true').lower() == 'true'
    find_tracking = request.args.get('tracking', 'true').lower() == 'true'

    def err(msg):
        yield f"data: {msg}\n\n"
        yield "data: __FAILURE__\n\n"

    if not url or not url.startswith('http'):
        return Response(stream_with_context(err('ERROR: Invalid or missing URL')), mimetype='text/event-stream')

    USER_AGENTS = [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15',
        'Mozilla/5.0 (X11; Linux x86_64; rv:120.0) Gecko/20100101 Firefox/120.0',
    ]

    def generate():
        global _html_cache

        yield f"data: Fetching {url}\n\n"

        try:
            headers = {
                'User-Agent': random.choice(USER_AGENTS),
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.9',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1',
            }
            resp = req.get(url, headers=headers, timeout=10, allow_redirects=True)

            if resp.status_code != 200:
                yield f"data: ERROR: HTTP {resp.status_code} — {url}\n\n"
                yield "data: __FAILURE__\n\n"
                return

            content_type = resp.headers.get('Content-Type', '')
            if 'html' not in content_type:
                yield f"data: ERROR: Not an HTML page (Content-Type: {content_type})\n\n"
                yield "data: __FAILURE__\n\n"
                return

            raw_html = resp.text
            _html_cache[url] = extract_content_html(raw_html, url)
            soup = BeautifulSoup(raw_html, 'html.parser')

            title_tag  = soup.find('title')
            page_title = title_tag.get_text().strip()[:120] if title_tag else ''
            yield f"data: PAGEINFO: {url}|{page_title}\n\n"
            yield f"data: Page: {page_title or url}\n\n"

            found_pdfs   = set()
            found_sites  = set()
            found_links  = set()
            found_images = set()
            found_trackers = set()

            for a in soup.find_all('a', href=True):
                href = a['href'].strip()
                if not href or href.startswith('mailto:') or href.startswith('javascript:'):
                    continue
                full = urljoin(url, href).split('#')[0].rstrip('/')
                if not full.startswith('http'):
                    continue

                # PDF: check URL path extension AND data-file-name attribute (Finalsite CMS)
                data_file = (a.get('data-file-name') or '').strip()
                url_path  = full.lower().split('?')[0]
                is_pdf    = url_path.endswith('.pdf') or data_file.lower().endswith('.pdf')
                if find_pdf and is_pdf and full not in found_pdfs:
                    found_pdfs.add(full)
                    display = data_file or url_path.split('/')[-1] or full
                    yield f"data: PDF: {full}|{display}\n\n"
                    continue

                if find_sites and 'sites.google.com' in full and full not in found_sites:
                    found_sites.add(full)
                    yield f"data: SITE: {full}|{url}\n\n"
                    continue

                if full not in found_links:
                    found_links.add(full)
                    link_text = a.get_text().strip()[:80]
                    yield f"data: LINK: {full}|{link_text}\n\n"

            if find_images:
                for img in soup.find_all('img'):
                    src = (img.get('src') or img.get('data-src') or
                           img.get('data-lazy-src') or '').strip()
                    if src:
                        full_img = urljoin(url, src)
                        if full_img.startswith('http') and full_img not in found_images:
                            found_images.add(full_img)
                            yield f"data: IMG: {full_img}|{url}\n\n"
                for source in soup.find_all('source'):
                    for part in (source.get('srcset') or '').split(','):
                        src = part.strip().split()[0] if part.strip() else ''
                        if src:
                            full_img = urljoin(url, src)
                            if full_img.startswith('http') and full_img not in found_images:
                                found_images.add(full_img)
                                yield f"data: IMG: {full_img}|{url}\n\n"

            if find_tracking:
                script_text = ''
                for script in soup.find_all('script'):
                    script_text += ' ' + (script.get('src') or '')
                    script_text += ' ' + (script.get_text() or '')
                for name, patterns in TRACKER_PATTERNS:
                    for pattern in patterns:
                        if pattern.lower() in script_text.lower():
                            if name not in found_trackers:
                                found_trackers.add(name)
                                yield f"data: TRACKER: {name}|{url}\n\n"
                            break

            yield f"data: \n\n"
            yield (
                f"data: Done. {len(found_links)} link(s), {len(found_pdfs)} PDF(s), "
                f"{len(found_sites)} Google Sites, {len(found_images)} image(s), "
                f"{len(found_trackers)} tracker(s) found.\n\n"
            )
            yield "data: __SUCCESS__\n\n"

        except Exception as e:
            yield f"data: Error: {e}\n\n"
            yield "data: __FAILURE__\n\n"

    resp = Response(stream_with_context(generate()), mimetype='text/event-stream')
    resp.headers['Cache-Control'] = 'no-cache'
    resp.headers['X-Accel-Buffering'] = 'no'
    return resp


@app.route('/api/broken-link-check')
def run_broken_link_check():
    import requests as req
    from bs4 import BeautifulSoup
    from urllib.parse import urlparse, urljoin
    from collections import deque
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import time, random, json

    url          = request.args.get('url', '').strip()
    mode         = request.args.get('mode', 'page')   # 'page' | 'site'
    check_imgs   = request.args.get('images',   'true').lower()  == 'true'
    check_pdfs   = request.args.get('pdfs',     'true').lower()  == 'true'
    check_ext    = request.args.get('external', 'false').lower() == 'true'
    check_social = request.args.get('social',   'false').lower() == 'true'

    SOCIAL_DOMAINS = [
        'facebook.com', 'fb.com', 'fbcdn.net',
        'twitter.com', 'x.com', 't.co',
        'instagram.com', 'instagr.am',
        'linkedin.com', 'lnkd.in',
        'youtube.com', 'youtu.be', 'yt.be',
        'tiktok.com', 'vm.tiktok.com',
        'pinterest.com', 'pin.it',
        'snapchat.com', 'snap.com',
        'reddit.com', 'redd.it',
        'tumblr.com', 'vimeo.com',
        'flickr.com', 'threads.net',
    ]

    def is_social_url(u):
        netloc = urlparse(u).netloc.lower().removeprefix('www.')
        return any(netloc == s or netloc.endswith('.' + s) for s in SOCIAL_DOMAINS)

    def err(msg):
        yield f"data: {msg}\n\n"
        yield "data: __FAILURE__\n\n"

    if not url or not url.startswith('http'):
        return Response(stream_with_context(err('ERROR: Invalid or missing URL')), mimetype='text/event-stream')

    USER_AGENTS = [
        # Chrome — Windows
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        # Chrome — Mac
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
        # Chrome — Linux
        'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
        # Firefox — Windows
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:122.0) Gecko/20100101 Firefox/122.0',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:120.0) Gecko/20100101 Firefox/120.0',
        # Firefox — Mac / Linux
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:124.0) Gecko/20100101 Firefox/124.0',
        'Mozilla/5.0 (X11; Linux x86_64; rv:124.0) Gecko/20100101 Firefox/124.0',
        # Safari — Mac
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4_1) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4.1 Safari/605.1.15',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15',
        # Edge — Windows
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0',
        # Mobile
        'Mozilla/5.0 (Linux; Android 14; Pixel 8) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Mobile Safari/537.36',
        'Mozilla/5.0 (iPhone; CPU iPhone OS 17_4_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4.1 Mobile/15E148 Safari/604.1',
    ]

    def make_headers():
        ua = random.choice(USER_AGENTS)
        return {
            'User-Agent': ua,
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.9',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        }

    def check_url_status(check_u):
        # Use GET with stream=True (never reads body, just headers) rather than HEAD.
        # HEAD is unreliable — many servers return 403/404 for HEAD but serve fine on GET.
        def _get(hdrs):
            r = req.get(check_u, headers=hdrs, timeout=7, allow_redirects=True, stream=True)
            r.close()
            return r

        try:
            r = _get(make_headers())

            # Retry once on 429, honouring Retry-After if present
            if r.status_code == 429:
                wait = min(int(r.headers.get('Retry-After', 3)), 15)
                time.sleep(wait + random.uniform(0.5, 2.0))
                r = _get(make_headers())

            return r.status_code, r.reason or str(r.status_code)
        except req.exceptions.Timeout:
            return 0, 'Timeout'
        except req.exceptions.SSLError:
            return 0, 'SSL Error'
        except req.exceptions.TooManyRedirects:
            return 0, 'Too Many Redirects'
        except req.exceptions.InvalidURL:
            return 0, 'Invalid URL'
        except req.exceptions.ConnectionError:
            return 0, 'Connection Error'
        except Exception as e:
            return 0, type(e).__name__

    def get_snippet(element):
        try:
            html = str(element)
            return html[:600] + ('…' if len(html) > 600 else '')
        except Exception:
            return ''

    def generate():
        try:
            parsed_start     = urlparse(url)
            base_domain      = parsed_start.netloc
            base_domain_norm = base_domain.removeprefix('www.')

            visited_pages = set()
            queued_pages  = {url}
            checked_urls  = set()
            queue = deque([url])

            pages_crawled = 0
            total_checked = 0
            broken_count  = 0

            yield f"data: Starting broken link check — {url}\n\n"

            while queue:
                current = queue.popleft()
                if current in visited_pages:
                    continue
                visited_pages.add(current)
                pages_crawled += 1

                yield f"data: PAGE: {pages_crawled}\n\n"
                yield f"data: Scanning [{pages_crawled}]: {current}\n\n"

                try:
                    resp = req.get(current, headers=make_headers(), timeout=12, allow_redirects=True)
                    if resp.status_code != 200:
                        yield f"data: Skipped (HTTP {resp.status_code}): {current}\n\n"
                        continue
                    if 'html' not in resp.headers.get('Content-Type', ''):
                        continue

                    soup = BeautifulSoup(resp.text, 'html.parser')

                    # Queue internal pages for crawling (site mode)
                    if mode == 'site':
                        for a in soup.find_all('a', href=True):
                            href = a['href'].strip()
                            if not href or href.startswith(('mailto:', 'javascript:', '#', 'tel:')):
                                continue
                            full = urljoin(current, href).split('#')[0].rstrip('/')
                            if not full.startswith('http'):
                                continue
                            if urlparse(full).netloc.removeprefix('www.') == base_domain_norm and full not in visited_pages and full not in queued_pages:
                                queued_pages.add(full)
                                queue.append(full)

                    # Collect resources to check from this page
                    resources = []

                    for a in soup.find_all('a', href=True):
                        href = a['href'].strip()
                        if not href or href.startswith(('mailto:', 'javascript:', '#', 'tel:')):
                            continue
                        full = urljoin(current, href).split('#')[0].rstrip('/')
                        if not full.startswith('http'):
                            continue
                        if full in checked_urls:
                            continue

                        is_internal = urlparse(full).netloc.removeprefix('www.') == base_domain_norm
                        social      = is_social_url(full)
                        url_path    = full.lower().split('?')[0]
                        is_pdf      = url_path.endswith('.pdf')

                        if not is_internal:
                            if social and not check_social:
                                continue
                            # PDFs obey check_pdfs regardless of check_ext
                            if is_pdf and not check_pdfs:
                                continue
                            if not is_pdf and not check_ext:
                                continue

                        if is_pdf:
                            if check_pdfs:
                                resources.append(('pdf', full, a, current, social))
                        else:
                            resources.append(('link', full, a, current, social))

                    if check_imgs:
                        for img in soup.find_all('img'):
                            src = (img.get('src') or img.get('data-src') or img.get('data-lazy-src') or '').strip()
                            if not src:
                                continue
                            full_img = urljoin(current, src)
                            if not full_img.startswith('http') or full_img in checked_urls:
                                continue
                            resources.append(('image', full_img, img, current, False))

                    # Dedupe before checking (tuple index 1 is the URL)
                    new_resources = []
                    for item in resources:
                        if item[1] not in checked_urls:
                            checked_urls.add(item[1])
                            new_resources.append(item)

                    if new_resources:
                        yield f"data: Checking {len(new_resources)} resource(s)…\n\n"

                    check_results = []
                    try:
                        with ThreadPoolExecutor(max_workers=3) as executor:
                            future_map = {
                                executor.submit(check_url_status, r[1]): r
                                for r in new_resources
                            }
                            for fut in as_completed(future_map):
                                try:
                                    status, status_text = fut.result()
                                except Exception as fe:
                                    status, status_text = 0, type(fe).__name__
                                item = future_map[fut]
                                total_checked += 1
                                yield f"data: CHECKING: {total_checked}|{item[1]}\n\n"
                                check_results.append((item, status, status_text))
                    except Exception as te:
                        yield f"data: Warning: thread pool error — {te}\n\n"

                    for res, status, status_text in check_results:
                        res_type, res_url, element, source_page, is_social = res
                        if status == 0 or status >= 400:
                            # Social media: only 404 is a genuine broken link;
                            # 400/403/429 etc. are normal bot-blocking responses.
                            if is_social and status != 404:
                                continue
                            broken_count += 1
                            try:
                                payload = json.dumps({
                                    'type':       res_type,
                                    'url':        res_url,
                                    'status':     status,
                                    'statusText': status_text,
                                    'source':     source_page,
                                    'snippet':    get_snippet(element),
                                }, ensure_ascii=True)
                            except Exception:
                                payload = json.dumps({
                                    'type': res_type, 'url': res_url,
                                    'status': status, 'statusText': status_text,
                                    'source': source_page, 'snippet': '',
                                })
                            yield f"data: BROKEN: {payload}\n\n"
                            yield f"data: BROKEN_COUNT: {broken_count}\n\n"

                except Exception as e:
                    yield f"data: Error scanning {current}: {type(e).__name__}: {e}\n\n"

                time.sleep(0.1)

            yield f"data: \n\n"
            yield (
                f"data: Done. {pages_crawled} page(s) scanned, "
                f"{total_checked} resource(s) checked, "
                f"{broken_count} broken found.\n\n"
            )
            yield "data: __SUCCESS__\n\n"

        except GeneratorExit:
            return
        except Exception as e:
            yield f"data: Fatal error: {type(e).__name__}: {e}\n\n"
            yield "data: __FAILURE__\n\n"

    resp = Response(stream_with_context(generate()), mimetype='text/event-stream')
    resp.headers['Cache-Control'] = 'no-cache'
    resp.headers['X-Accel-Buffering'] = 'no'
    return resp


def _ensure_spell_ignore_table(conn):
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS spell_ignore (
                id       INT AUTO_INCREMENT PRIMARY KEY,
                word     VARCHAR(200) NOT NULL,
                added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY uq_word (word)
            )
        """)
        conn.commit()
        cur.close()
    except Exception:
        pass


@app.route('/api/spell-ignore', methods=['GET'])
def get_spell_ignore():
    conn = get_db_connection()
    if not conn:
        return jsonify({'words': []})
    try:
        _ensure_spell_ignore_table(conn)
        cur = conn.cursor()
        cur.execute("SELECT word, added_at FROM spell_ignore ORDER BY word ASC")
        words = [{'word': row[0], 'added_at': str(row[1])} for row in cur.fetchall()]
        cur.close()
        return jsonify({'words': words})
    except Exception as e:
        return jsonify({'words': [], 'error': str(e)})
    finally:
        conn.close()


@app.route('/api/spell-ignore', methods=['POST'])
def add_spell_ignore():
    word = ((request.get_json(force=True) or {}).get('word') or '').strip().lower()
    if not word:
        return jsonify({'ok': False, 'error': 'No word provided'}), 400
    conn = get_db_connection()
    if not conn:
        return jsonify({'ok': False, 'error': 'DB unavailable'}), 500
    try:
        _ensure_spell_ignore_table(conn)
        cur = conn.cursor()
        cur.execute("INSERT IGNORE INTO spell_ignore (word) VALUES (%s)", (word,))
        conn.commit()
        cur.close()
        return jsonify({'ok': True, 'word': word})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/spell-ignore/<path:word>', methods=['DELETE'])
def delete_spell_ignore(word):
    word = word.strip().lower()
    conn = get_db_connection()
    if not conn:
        return jsonify({'ok': False, 'error': 'DB unavailable'}), 500
    try:
        _ensure_spell_ignore_table(conn)
        cur = conn.cursor()
        cur.execute("DELETE FROM spell_ignore WHERE word = %s", (word,))
        conn.commit()
        cur.close()
        return jsonify({'ok': True, 'word': word})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/ollama-models')
def get_ollama_models_route():
    base_url = get_ollama_url()
    running, installed = list_ollama_models(base_url)
    default = running[0] if running else (installed[0] if installed else None)
    return jsonify({
        'ok': bool(installed or running),
        'running': running,
        'installed': installed,
        'default': default,
    })


@app.route('/api/seo-check')
def run_seo_check():
    import requests as req
    from bs4 import BeautifulSoup
    from urllib.parse import urlparse, urljoin
    from collections import deque
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import time, random, json

    url            = request.args.get('url', '').strip()
    mode           = request.args.get('mode', 'site')   # 'page' | 'site'
    find_pdf       = request.args.get('pdf', 'true').lower() == 'true'
    find_sites     = request.args.get('sites', 'true').lower() == 'true'
    find_external  = request.args.get('external', 'true').lower() == 'true'
    find_tracking  = request.args.get('tracking', 'true').lower() == 'true'
    find_alt       = request.args.get('alt', 'true').lower() == 'true'
    find_spell     = request.args.get('spell', 'true').lower() == 'true'
    find_ai        = request.args.get('ai', 'true').lower() == 'true'
    find_keywords  = request.args.get('kw', 'false').lower() == 'true'
    requested_model = (request.args.get('model') or '').strip()
    ollama_url     = get_ollama_url()
    OLLAMA_TIMEOUT = int(os.getenv('OLLAMA_TIMEOUT', '300'))
    check_ext_stat = request.args.get('check_external', 'false').lower() == 'true'
    check_social   = request.args.get('social', 'false').lower() == 'true'

    SOCIAL_DOMAINS = [
        'facebook.com', 'fb.com', 'fbcdn.net',
        'twitter.com', 'x.com', 't.co',
        'instagram.com', 'instagr.am',
        'linkedin.com', 'lnkd.in',
        'youtube.com', 'youtu.be', 'yt.be',
        'tiktok.com', 'vm.tiktok.com',
        'pinterest.com', 'pin.it',
        'snapchat.com', 'snap.com',
        'reddit.com', 'redd.it',
        'tumblr.com', 'vimeo.com',
        'flickr.com', 'threads.net',
    ]

    def is_social_url(u):
        netloc = urlparse(u).netloc.lower().removeprefix('www.')
        return any(netloc == s or netloc.endswith('.' + s) for s in SOCIAL_DOMAINS)

    def err(msg):
        yield f"data: {msg}\n\n"
        yield "data: __FAILURE__\n\n"

    if not url or not url.startswith('http'):
        return Response(stream_with_context(err('ERROR: Invalid or missing URL')), mimetype='text/event-stream')

    USER_AGENTS = [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15',
        'Mozilla/5.0 (X11; Linux x86_64; rv:120.0) Gecko/20100101 Firefox/120.0',
    ]

    def make_headers():
        return {
            'User-Agent': random.choice(USER_AGENTS),
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.9',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        }

    def check_url_status(check_u):
        def _get(hdrs):
            r = req.get(check_u, headers=hdrs, timeout=7, allow_redirects=True, stream=True)
            r.close()
            return r
        try:
            r = _get(make_headers())
            if r.status_code == 429:
                wait = min(int(r.headers.get('Retry-After', 3)), 15)
                time.sleep(wait + random.uniform(0.5, 2.0))
                r = _get(make_headers())
            return r.status_code, r.reason or str(r.status_code)
        except req.exceptions.Timeout:
            return 0, 'Timeout'
        except req.exceptions.SSLError:
            return 0, 'SSL Error'
        except req.exceptions.TooManyRedirects:
            return 0, 'Too Many Redirects'
        except req.exceptions.InvalidURL:
            return 0, 'Invalid URL'
        except req.exceptions.ConnectionError:
            return 0, 'Connection Error'
        except Exception as e:
            return 0, type(e).__name__

    def get_snippet(element):
        try:
            html = str(element)
            return html[:600] + ('…' if len(html) > 600 else '')
        except Exception:
            return ''

    def resolve_ollama_model():
        if requested_model:
            return requested_model
        running, installed = list_ollama_models(ollama_url)
        if running:
            return running[0]
        if installed:
            return installed[0]
        return None

    def get_ollama_suggestion(model, current_title, current_description, content_text):
        prompt = (
            "You are an SEO specialist. Based on the webpage content below, write an "
            "improved, accurate SEO title tag (max 60 characters) and meta description "
            "(max 155 characters) for this exact page. Respond with exactly these two "
            "lines and nothing else:\n"
            "TITLE: <suggested title>\n"
            "DESCRIPTION: <suggested description>\n\n"
            f"Current title: {current_title or '(none)'}\n"
            f"Current meta description: {current_description or '(none)'}\n\n"
            f"Page content:\n{content_text[:3000]}"
        )
        r = req.post(
            f"{ollama_url}/api/generate",
            json={'model': model, 'prompt': prompt, 'stream': False},
            timeout=OLLAMA_TIMEOUT,
        )
        r.raise_for_status()
        text = r.json().get('response', '')
        title_m = re.search(r'TITLE:\s*(.+)', text)
        desc_m  = re.search(r'DESCRIPTION:\s*(.+)', text, re.S)
        suggested_title = title_m.group(1).strip().strip('"') if title_m else ''
        suggested_desc  = desc_m.group(1).strip().strip('"').split('\n')[0] if desc_m else ''
        return suggested_title, suggested_desc

    def get_ollama_keywords(model, content_text, page_url):
        prompt = (
            "You are an expert SEO keyword strategist. Analyze the webpage content below and generate "
            "high-traffic, brand-specific keyword phrases for this site. This tool is used on ALL types "
            "of websites — agencies, retailers, restaurants, law firms, schools, nonprofits, SaaS products, "
            "medical practices, contractors, blogs, etc. Adapt your output to whatever industry this site is in.\n\n"
            f"Page URL: {page_url}\n\n"
            "--- INSTRUCTIONS ---\n\n"
            "STEP 1: Read the content and URL. Identify:\n"
            "  - The brand/organization name as it is commonly known (org_name)\n"
            "  - Any short version or abbreviation (org_short) — use the full name if no abbreviation exists\n"
            "  - The industry or business type (e.g., digital marketing agency, Italian restaurant, personal injury law firm)\n\n"
            "STEP 2: List the top 3 topics, products, or services this specific page focuses on.\n\n"
            "STEP 3: Generate 4–5 short branded keyword phrases (2–5 words each) this page should rank for.\n"
            "  - Every phrase must contain the brand name or abbreviation from Step 1\n"
            "  - Examples of the FORMAT (not the content — use the actual brand you found):\n"
            "      [Brand] [service], [Brand] [city], [Brand] [product] reviews, [Brand] vs competitors\n"
            "  - Add a location only if the site serves a specific geographic area\n"
            "  - No generic phrases — each phrase must only make sense for this specific brand\n\n"
            "STEP 4: Generate 6–9 long-tail phrases a real visitor would type into Google to reach this page.\n"
            "  - Every phrase must contain the brand name or abbreviation\n"
            "  - Match the phrasing to what someone in this industry's audience would search\n"
            "  - Cover a mix of: finding the brand (Navigational), learning something (Informational), "
            "taking action (Action-oriented)\n"
            "  - No phrase should be so generic it could apply to any company\n\n"
            "STEP 5: Label each long-tail phrase with its user intent: Navigational, Informational, or Action-oriented.\n\n"
            "--- OUTPUT FORMAT ---\n\n"
            "Return ONLY a valid JSON object. No markdown, no explanation, no code fences.\n\n"
            "{\n"
            "  \"org_name\": \"<full brand name detected from content>\",\n"
            "  \"org_short\": \"<short name or same as org_name>\",\n"
            "  \"current_topics\": [\"<topic 1>\", \"<topic 2>\", \"<topic 3>\"],\n"
            "  \"target_keywords\": [\n"
            "    {\"phrase\": \"<brand short keyword 1>\"},\n"
            "    {\"phrase\": \"<brand short keyword 2>\"},\n"
            "    {\"phrase\": \"<brand short keyword 3>\"},\n"
            "    {\"phrase\": \"<brand short keyword 4>\"},\n"
            "    {\"phrase\": \"<brand short keyword 5>\"}\n"
            "  ],\n"
            "  \"community_search_suggestions\": [\n"
            "    {\"phrase\": \"<long-tail phrase 1>\", \"user_intent\": \"<intent>\"},\n"
            "    {\"phrase\": \"<long-tail phrase 2>\", \"user_intent\": \"<intent>\"},\n"
            "    {\"phrase\": \"<long-tail phrase 3>\", \"user_intent\": \"<intent>\"},\n"
            "    {\"phrase\": \"<long-tail phrase 4>\", \"user_intent\": \"<intent>\"},\n"
            "    {\"phrase\": \"<long-tail phrase 5>\", \"user_intent\": \"<intent>\"},\n"
            "    {\"phrase\": \"<long-tail phrase 6>\", \"user_intent\": \"<intent>\"}\n"
            "  ]\n"
            "}\n\n"
            "--- WEBPAGE CONTENT ---\n"
            f"{content_text[:4000]}"
        )
        r = req.post(
            f"{ollama_url}/api/generate",
            json={'model': model, 'prompt': prompt, 'stream': False},
            timeout=OLLAMA_TIMEOUT,
        )
        r.raise_for_status()
        text = r.json().get('response', '').strip()
        # Strip markdown code fences that some models emit
        text = re.sub(r'^```(?:json)?\s*', '', text)
        text = re.sub(r'\s*```$', '', text).strip()
        # Extract the outermost JSON object
        m = re.search(r'\{[\s\S]*\}', text)
        raw = m.group() if m else text
        # Remove trailing commas before ] or } which some models produce
        raw = re.sub(r',\s*([}\]])', r'\1', raw)
        return json.loads(raw)

    spell = get_spell_checker() if find_spell else None

    def generate():
        global _html_cache
        _html_cache = {}  # clear previous scrape

        parsed_start  = urlparse(url)
        base_domain   = parsed_start.netloc.lower().removeprefix('www.')
        is_ccsd_domain = 'ccsd.net' in base_domain

        visited         = set()
        queued          = {url}
        checked_urls    = set()
        queue           = deque([url])

        found_pdfs      = set()
        found_sites     = set()
        found_externals = set()
        found_trackers  = set()
        found_images    = set()
        crawled_pages   = []  # [{'url':..., 'title':..., 'description':...}] — used for the spellcheck/AI passes that run after the crawl

        pages_crawled = 0
        total_checked = 0
        broken_count  = 0
        spell_count   = 0
        ai_count      = 0
        keyword_count = 0

        yield f"data: Starting SEO check — {url}\n\n"
        yield f"data: Domain: {base_domain}\n\n"

        while queue:
            current = queue.popleft()
            if current in visited:
                continue
            visited.add(current)

            try:
                resp = req.get(current, headers=make_headers(), timeout=12, allow_redirects=True)
                if resp.status_code != 200:
                    yield f"data: Skipped (HTTP {resp.status_code}): {current}\n\n"
                    continue
                if 'html' not in resp.headers.get('Content-Type', ''):
                    continue

                # /fs/pages/ shortlink dedup — ccsd.net only.
                # These CMS shortlinks redirect to canonical URLs; if the destination
                # was already crawled (or will be), skip the /fs/pages/ duplicate.
                effective = current
                if is_ccsd_domain and re.search(r'/fs/pages/\d+', current):
                    canonical = resp.url.rstrip('/')
                    if (canonical != current and
                            urlparse(canonical).netloc.lower().removeprefix('www.') == base_domain):
                        if canonical in visited:
                            yield f"data: Deduped: {current} → {canonical}\n\n"
                            continue
                        visited.add(canonical)
                        queued.add(canonical)
                        effective = canonical

                pages_crawled += 1
                yield f"data: PAGE: {pages_crawled}\n\n"
                yield f"data: [{pages_crawled}] {effective}\n\n"

                raw_html     = resp.text
                content_html = extract_content_html(raw_html, effective)
                _html_cache[effective] = content_html
                soup = BeautifulSoup(raw_html, 'html.parser')

                # Title & meta description
                title_tag   = soup.find('title')
                page_title  = title_tag.get_text().strip()[:160] if title_tag else ''
                desc_tag    = soup.find('meta', attrs={'name': re.compile(r'^description$', re.I)})
                description = (desc_tag.get('content') or '').strip()[:300] if desc_tag else ''
                yield f"data: PAGEINFO: {json.dumps({'url': effective, 'title': page_title, 'description': description}, ensure_ascii=True)}\n\n"
                crawled_pages.append({'url': effective, 'title': page_title, 'description': description})

                pending_checks = []  # (type, url, element, source_page, is_social)

                # Walk <a> tags: PDFs / Google Sites / External links / internal queueing / broken-check collection
                for a in soup.find_all('a', href=True):
                    href = a['href'].strip()
                    if not href or href.startswith(('mailto:', 'javascript:', '#', 'tel:')):
                        continue
                    full = urljoin(effective, href).split('#')[0].rstrip('/')
                    if not full.startswith('http'):
                        continue

                    data_file = (a.get('data-file-name') or '').strip()
                    url_path  = full.lower().split('?')[0]
                    is_pdf    = url_path.endswith('.pdf') or data_file.lower().endswith('.pdf')

                    if find_pdf and is_pdf and full not in found_pdfs:
                        found_pdfs.add(full)
                        yield f"data: PDF: {full}|{effective}\n\n"

                    if find_sites and 'sites.google.com' in full and full not in found_sites:
                        found_sites.add(full)
                        yield f"data: SITE: {full}|{effective}\n\n"

                    link_domain = urlparse(full).netloc.lower().removeprefix('www.')
                    is_internal = link_domain == base_domain

                    if is_internal:
                        if mode == 'site' and full not in visited and full not in queued:
                            queued.add(full)
                            queue.append(full)
                    elif find_external and full not in found_externals:
                        found_externals.add(full)
                        rel = a.get('rel') or []
                        if isinstance(rel, str):
                            rel = rel.split()
                        nofollow = 'nofollow' in [r.lower() for r in rel]
                        is_ccsd  = link_domain == 'ccsd.net' or link_domain.endswith('.ccsd.net')
                        yield f"data: EXTERNAL: {full}|{effective}|{1 if nofollow else 0}|{1 if is_ccsd else 0}\n\n"

                    if full not in checked_urls:
                        social = is_social_url(full)
                        if is_pdf:
                            if find_pdf:
                                pending_checks.append(('pdf', full, a, effective, social))
                        elif is_internal:
                            pending_checks.append(('link', full, a, effective, social))
                        elif check_ext_stat and not (social and not check_social):
                            pending_checks.append(('link', full, a, effective, social))

                # Images: alt-tag analysis + broken-check collection
                if find_alt:
                    for img in soup.find_all('img'):
                        src = (img.get('src') or img.get('data-src') or img.get('data-lazy-src') or '').strip()
                        if not src:
                            continue
                        full_img = urljoin(effective, src)
                        if not full_img.startswith('http'):
                            continue
                        alt     = img.get('alt')
                        has_alt = alt is not None and alt.strip() != ''
                        if full_img not in found_images:
                            found_images.add(full_img)
                            payload = json.dumps({
                                'src': full_img, 'alt': alt or '', 'hasAlt': has_alt, 'source': effective,
                            }, ensure_ascii=True)
                            yield f"data: IMGALT: {payload}\n\n"
                        if full_img not in checked_urls:
                            pending_checks.append(('image', full_img, img, effective, False))

                # Tracking/analytics scripts
                if find_tracking:
                    script_text = ''
                    for script in soup.find_all('script'):
                        script_text += ' ' + (script.get('src') or '')
                        script_text += ' ' + (script.get_text() or '')
                    for name, patterns in TRACKER_PATTERNS:
                        for pattern in patterns:
                            if pattern.lower() in script_text.lower():
                                key = f"{name}|{effective}"
                                if key not in found_trackers:
                                    found_trackers.add(key)
                                    yield f"data: TRACKER: {name}|{effective}\n\n"
                                break

                # Broken-link/image/pdf status checks
                new_checks = []
                for item in pending_checks:
                    if item[1] not in checked_urls:
                        checked_urls.add(item[1])
                        new_checks.append(item)

                if new_checks:
                    yield f"data: Checking {len(new_checks)} resource(s)…\n\n"

                check_results = []
                try:
                    with ThreadPoolExecutor(max_workers=3) as executor:
                        future_map = {executor.submit(check_url_status, r[1]): r for r in new_checks}
                        for fut in as_completed(future_map):
                            try:
                                status, status_text = fut.result()
                            except Exception as fe:
                                status, status_text = 0, type(fe).__name__
                            item = future_map[fut]
                            total_checked += 1
                            yield f"data: CHECKING: {total_checked}|{item[1]}\n\n"
                            check_results.append((item, status, status_text))
                except Exception as te:
                    yield f"data: Warning: thread pool error — {te}\n\n"

                for res, status, status_text in check_results:
                    res_type, res_url, element, source_page, social = res
                    if status == 0 or status >= 400:
                        if social and status != 404:
                            continue
                        broken_count += 1
                        try:
                            payload = json.dumps({
                                'type': res_type, 'url': res_url, 'status': status,
                                'statusText': status_text, 'source': source_page,
                                'snippet': get_snippet(element),
                            }, ensure_ascii=True)
                        except Exception:
                            payload = json.dumps({
                                'type': res_type, 'url': res_url, 'status': status,
                                'statusText': status_text, 'source': source_page, 'snippet': '',
                            })
                        yield f"data: BROKEN: {payload}\n\n"
                        yield f"data: BROKEN_COUNT: {broken_count}\n\n"

            except Exception as e:
                yield f"data: Error ({current}): {type(e).__name__}: {e}\n\n"

            time.sleep(0.15)

        # ---- Phase 1.5: PDF page count ----
        if find_pdf and found_pdfs:
            yield f"data: Counting PDF pages ({len(found_pdfs)} files)…\n\n"

            def count_pdf_pages(pdf_url):
                try:
                    r = req.get(pdf_url, headers={'User-Agent': random.choice(USER_AGENTS)},
                                timeout=20, stream=True)
                    if r.status_code != 200:
                        r.close()
                        return None
                    data = b''
                    for chunk in r.iter_content(8192):
                        data += chunk
                        if len(data) >= 524288:  # 512 KB is enough for page-tree metadata
                            break
                    r.close()
                    counts = re.findall(rb'/Count\s+(\d+)', data)
                    if counts:
                        return max(int(c) for c in counts)
                    pages = len(re.findall(rb'/Type\s*/Page\b', data))
                    return pages if pages > 0 else None
                except Exception:
                    return None

            with ThreadPoolExecutor(max_workers=3) as pdf_ex:
                pdf_futures = {pdf_ex.submit(count_pdf_pages, u): u for u in found_pdfs}
                for fut in as_completed(pdf_futures):
                    pdf_url = pdf_futures[fut]
                    try:
                        pages = fut.result()
                    except Exception:
                        pages = None
                    if pages is not None:
                        difficulty = 'easy' if pages == 1 else 'hard'
                    else:
                        difficulty = 'unknown'
                    yield f"data: PDF_INFO: {json.dumps({'url': pdf_url, 'pages': pages, 'difficulty': difficulty}, ensure_ascii=True)}\n\n"

        # ---- Phase 2: spell-check every crawled page, now that discovery/link
        # checking is fully done (keeps slow per-page work from stalling the crawl) ----
        if find_spell and crawled_pages:
            # Load ignore list from DB once before the loop
            ignore_words = set()
            try:
                ig_conn = get_db_connection()
                if ig_conn:
                    _ensure_spell_ignore_table(ig_conn)
                    ig_cur = ig_conn.cursor()
                    ig_cur.execute("SELECT word FROM spell_ignore")
                    ignore_words = {row[0].lower() for row in ig_cur.fetchall()}
                    ig_cur.close()
                    ig_conn.close()
            except Exception:
                pass

            yield f"data: \n\n"
            yield f"data: Spell-checking {len(crawled_pages)} page(s)... ({len(ignore_words)} word(s) in ignore list)\n\n"
            for pdata in crawled_pages:
                page_url = pdata['url']
                content_html = _html_cache.get(page_url, '')
                if not content_html:
                    continue
                content_soup = BeautifulSoup(content_html, 'html.parser')
                candidates = []
                for node in content_soup.find_all(string=True):
                    if node.parent and node.parent.name in ('script', 'style'):
                        continue
                    for w in re.findall(r"[A-Za-z']+", str(node)):
                        if len(w) < 3 or w[:1].isupper():
                            continue
                        candidates.append(w.strip("'").lower())

                unknown = spell.unknown(candidates) if candidates else set()
                unknown -= ignore_words  # filter out ignored words
                if not unknown:
                    continue
                reported = set()
                for node in content_soup.find_all(string=True):
                    if node.parent and node.parent.name in ('script', 'style'):
                        continue
                    for w in re.findall(r"[A-Za-z']+", str(node)):
                        if len(w) < 3 or w[:1].isupper():
                            continue
                        wl = w.strip("'").lower()
                        if wl in unknown and wl not in reported:
                            reported.add(wl)
                            suggestion = spell.correction(wl) or ''
                            snippet = get_snippet(node.parent) if node.parent else str(node).strip()[:300]
                            spell_count += 1
                            payload = json.dumps({
                                'word': w, 'suggestion': suggestion,
                                'snippet': snippet, 'source': page_url,
                            }, ensure_ascii=True)
                            yield f"data: SPELL: {payload}\n\n"
                            yield f"data: SPELL_COUNT: {spell_count}\n\n"

        # ---- Phase 3: ask Ollama for title/description suggestions for every
        # crawled page, now that everything else has finished ----
        if find_ai and crawled_pages:
            resolved_model = resolve_ollama_model()
            if resolved_model:
                yield f"data: \n\n"
                yield f"data: Generating AI suggestions for {len(crawled_pages)} page(s) using {resolved_model}...\n\n"
                for pdata in crawled_pages:
                    page_url = pdata['url']
                    content_html = _html_cache.get(page_url, '')
                    if not content_html:
                        continue
                    try:
                        content_text = BeautifulSoup(content_html, 'html.parser').get_text(separator=' ', strip=True)
                        content_text = re.sub(r'\s+', ' ', content_text).strip()
                        if not content_text:
                            continue
                        yield f"data: Asking Ollama ({resolved_model}) for suggestions — {page_url}\n\n"
                        sugg_title, sugg_desc = get_ollama_suggestion(
                            resolved_model, pdata['title'], pdata['description'], content_text,
                        )
                        if sugg_title or sugg_desc:
                            ai_count += 1
                            payload = json.dumps({
                                'url': page_url,
                                'currentTitle': pdata['title'],
                                'currentDescription': pdata['description'],
                                'suggestedTitle': sugg_title,
                                'suggestedDescription': sugg_desc,
                            }, ensure_ascii=True)
                            yield f"data: AISEO: {payload}\n\n"
                    except req.exceptions.ReadTimeout:
                        yield f"data: AI suggestion timed out for {page_url} (>{OLLAMA_TIMEOUT}s) — skipping.\n\n"
                    except req.exceptions.ConnectionError:
                        yield f"data: AI suggestion skipped — could not connect to Ollama at {ollama_url}\n\n"
                    except Exception as e:
                        yield f"data: AI suggestion failed for {page_url}: {type(e).__name__}: {e}\n\n"
            else:
                yield f"data: AI suggestions disabled — no Ollama model found at {ollama_url}\n\n"

        # ---- Phase 4: keyword analysis (Ollama) + Google Trends interest scores ----
        if find_keywords and crawled_pages:
            resolved_model = resolve_ollama_model()
            if resolved_model:
                yield f"data: \n\n"
                yield f"data: Analyzing keywords for {len(crawled_pages)} page(s) using {resolved_model}...\n\n"
                trends_cache = {}
                for pdata in crawled_pages:
                    page_url     = pdata['url']
                    content_html = _html_cache.get(page_url, '')
                    if not content_html:
                        continue
                    try:
                        content_text = BeautifulSoup(content_html, 'html.parser').get_text(separator=' ', strip=True)
                        content_text = re.sub(r'\s+', ' ', content_text).strip()
                        if not content_text:
                            continue
                        yield f"data: Analyzing keywords — {page_url}\n\n"
                        kw_data        = get_ollama_keywords(resolved_model, content_text, page_url)
                        suggestions    = kw_data.get('community_search_suggestions', [])
                        target_kws     = kw_data.get('target_keywords', [])
                        all_phrases    = (
                            [t.get('phrase', '') for t in target_kws if t.get('phrase')] +
                            [s.get('phrase', '') for s in suggestions if s.get('phrase')]
                        )
                        new_phrases = [p for p in all_phrases if p not in trends_cache]
                        if new_phrases:
                            yield f"data: Checking Google Trends for {len(new_phrases)} phrase(s)...\n\n"
                            trends_cache.update(get_keyword_trends(new_phrases))
                        trends = {p: trends_cache.get(p) for p in all_phrases}
                        keyword_count += 1
                        payload = json.dumps({
                            'url':             page_url,
                            'org_name':        kw_data.get('org_name', ''),
                            'org_short':       kw_data.get('org_short', ''),
                            'topics':          kw_data.get('current_topics', []),
                            'target_keywords': target_kws,
                            'suggestions':     suggestions,
                            'trends':          trends,
                        }, ensure_ascii=True)
                        yield f"data: KEYWORDS: {payload}\n\n"
                    except req.exceptions.ReadTimeout:
                        yield f"data: Keyword analysis timed out for {page_url} — skipping.\n\n"
                    except req.exceptions.ConnectionError:
                        yield f"data: Keyword analysis skipped — could not connect to Ollama at {ollama_url}\n\n"
                    except Exception as e:
                        yield f"data: Keyword analysis failed for {page_url}: {type(e).__name__}: {e}\n\n"
            else:
                yield f"data: Keyword analysis disabled — no Ollama model found at {ollama_url}\n\n"

        yield f"data: \n\n"
        yield (
            f"data: Done. {pages_crawled} page(s) crawled, {total_checked} resource(s) checked. "
            f"{len(found_pdfs)} PDF(s), {len(found_sites)} Google Sites, {len(found_externals)} external link(s), "
            f"{len(found_trackers)} tracker event(s), {len(found_images)} image(s), "
            f"{spell_count} spelling issue(s), {ai_count} AI suggestion(s), "
            f"{keyword_count} keyword analysis page(s), "
            f"{broken_count} broken resource(s) found.\n\n"
        )
        yield "data: __SUCCESS__\n\n"

    resp = Response(stream_with_context(generate()), mimetype='text/event-stream')
    resp.headers['Cache-Control'] = 'no-cache'
    resp.headers['X-Accel-Buffering'] = 'no'
    return resp


@app.route('/api/page-html')
def get_page_html():
    page_url = request.args.get('url', '').strip()
    html = _html_cache.get(page_url, '')
    return jsonify({'html': html, 'found': bool(html)})


@app.route('/api/export/html-zip')
def export_html_zip():
    import zipfile
    from urllib.parse import urlparse

    if not _html_cache:
        return 'No HTML cached — run a scrape first', 400

    buf = io.BytesIO()
    used_names = set()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for page_url, content in _html_cache.items():
            name = url_to_filename(page_url)
            base, ext = name[:-4], name[-4:]
            final, n = name, 2
            while final in used_names:
                final = f"{base}_{n}{ext}"
                n += 1
            used_names.add(final)
            zf.writestr(final, clean_export_html(content) if content else '')
    buf.seek(0)

    first_host = urlparse(next(iter(_html_cache))).netloc.lower().removeprefix('www.') or 'site'
    zip_name = f"{first_host}_html.zip"

    return send_file(buf, mimetype='application/zip', as_attachment=True, download_name=zip_name)


@app.route('/api/export/images-zip', methods=['POST'])
def export_images_zip():
    import zipfile
    import requests as req
    from urllib.parse import urlparse

    data       = request.get_json(silent=True) or {}
    image_urls = data.get('images', [])
    if not image_urls:
        return 'No images to export', 400

    buf = io.BytesIO()
    used_names = set()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for img_url in image_urls:
            try:
                resp = req.get(img_url, timeout=15)
                if resp.status_code != 200:
                    continue
                filename = img_url.split('/')[-1].split('?')[0] or 'image'
                if '.' not in filename:
                    ct  = resp.headers.get('Content-Type', 'image/jpeg')
                    ext = ct.split('/')[-1].split(';')[0].strip() or 'jpg'
                    filename = f'{filename}.{ext}'
                filename = re.sub(r'[^a-zA-Z0-9\-_.]', '_', filename)

                final, n = filename, 2
                while final in used_names:
                    stem, dot, ext = filename.rpartition('.')
                    final = f"{stem}_{n}{dot}{ext}" if dot else f"{filename}_{n}"
                    n += 1
                used_names.add(final)
                zf.writestr(final, resp.content)
            except Exception:
                continue

    if not used_names:
        return 'Could not download any images', 502

    buf.seek(0)

    first_host = urlparse(image_urls[0]).netloc.lower().removeprefix('www.') or 'site'
    zip_name = f"{first_host}_images.zip"

    return send_file(buf, mimetype='application/zip', as_attachment=True, download_name=zip_name)


@app.route('/api/download-image')
def download_image():
    import requests as req
    img_url = request.args.get('url', '').strip()
    if not img_url.startswith('http'):
        return 'Invalid URL', 400
    try:
        resp = req.get(img_url, timeout=15, stream=True)
        filename = img_url.split('/')[-1].split('?')[0] or 'image'
        if not filename or '.' not in filename:
            ct = resp.headers.get('Content-Type', 'image/jpeg')
            ext = ct.split('/')[-1].split(';')[0].strip() or 'jpg'
            filename = f'image.{ext}'
        return Response(
            resp.iter_content(chunk_size=8192),
            content_type=resp.headers.get('Content-Type', 'image/jpeg'),
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        return str(e), 500


# ── PDF to Image ─────────────────────────────────────────────────────────────

@app.route('/pdf-to-image')
def pdf_to_image_page():
    return render_template('pdf_to_image.html')


@app.route('/api/pdf-scrape')
def run_pdf_scrape():
    """SSE: crawl a single page or an entire domain, reporting only PDF links found."""
    import requests as req
    from bs4 import BeautifulSoup
    from urllib.parse import urlparse, urljoin
    from collections import deque
    import time
    import random

    url  = request.args.get('url', '').strip()
    mode = request.args.get('mode', 'page')  # 'page' or 'site'

    def err(msg):
        yield f"data: {msg}\n\n"
        yield "data: __FAILURE__\n\n"

    if not url or not url.startswith('http'):
        return Response(stream_with_context(err('ERROR: Invalid or missing URL')), mimetype='text/event-stream')

    USER_AGENTS = [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15',
        'Mozilla/5.0 (X11; Linux x86_64; rv:120.0) Gecko/20100101 Firefox/120.0',
    ]

    def fetch_html(target):
        headers = {
            'User-Agent': random.choice(USER_AGENTS),
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.9',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        }
        return req.get(target, headers=headers, timeout=10, allow_redirects=True)

    def pdfs_on_page(soup, current):
        found = []
        for a in soup.find_all('a', href=True):
            href = a['href'].strip()
            if not href or href.startswith('mailto:') or href.startswith('javascript:'):
                continue
            full = urljoin(current, href).split('#')[0].rstrip('/')
            if not full.startswith('http'):
                continue
            data_file = (a.get('data-file-name') or '').strip()
            url_path  = full.lower().split('?')[0]
            is_pdf    = url_path.endswith('.pdf') or data_file.lower().endswith('.pdf')
            found.append((full, is_pdf))
        return found

    def generate():
        found_pdfs = set()

        if mode == 'site':
            parsed_start = urlparse(url)
            base_domain  = parsed_start.netloc.lower().removeprefix('www.')
            visited = set()
            queued  = {url}
            queue   = deque([url])
            pages_crawled = 0

            yield f"data: Starting crawl of {url}\n\n"
            yield f"data: Domain: {base_domain}\n\n"

            while queue:
                current = queue.popleft()
                if current in visited:
                    continue
                visited.add(current)
                pages_crawled += 1
                yield f"data: PAGE: {pages_crawled}\n\n"
                yield f"data: [{pages_crawled}] {current}\n\n"

                try:
                    resp = fetch_html(current)
                    if resp.status_code != 200:
                        yield f"data: Skipped (HTTP {resp.status_code}): {current}\n\n"
                        continue
                    content_type = resp.headers.get('Content-Type', '')
                    if 'html' not in content_type:
                        continue

                    soup = BeautifulSoup(resp.text, 'html.parser')

                    for full, is_pdf in pdfs_on_page(soup, current):
                        if is_pdf:
                            if full not in found_pdfs:
                                found_pdfs.add(full)
                                yield f"data: PDF: {full}|{current}\n\n"
                            continue

                        link_domain = urlparse(full).netloc.lower().removeprefix('www.')
                        if link_domain == base_domain and full not in visited and full not in queued:
                            queued.add(full)
                            queue.append(full)

                except Exception as e:
                    yield f"data: Error ({current}): {e}\n\n"

                time.sleep(0.15)

            yield f"data: \n\n"
            yield f"data: Done. {pages_crawled} page(s) crawled, {len(found_pdfs)} PDF(s) found.\n\n"
            yield "data: __SUCCESS__\n\n"

        else:
            yield f"data: Fetching {url}\n\n"
            try:
                resp = fetch_html(url)
                if resp.status_code != 200:
                    yield f"data: ERROR: HTTP {resp.status_code} — {url}\n\n"
                    yield "data: __FAILURE__\n\n"
                    return
                content_type = resp.headers.get('Content-Type', '')
                if 'html' not in content_type:
                    yield f"data: ERROR: Not an HTML page (Content-Type: {content_type})\n\n"
                    yield "data: __FAILURE__\n\n"
                    return

                soup = BeautifulSoup(resp.text, 'html.parser')
                for full, is_pdf in pdfs_on_page(soup, url):
                    if is_pdf and full not in found_pdfs:
                        found_pdfs.add(full)
                        yield f"data: PDF: {full}|{url}\n\n"

                yield f"data: \n\n"
                yield f"data: Done. {len(found_pdfs)} PDF(s) found.\n\n"
                yield "data: __SUCCESS__\n\n"
            except Exception as e:
                yield f"data: Error: {e}\n\n"
                yield "data: __FAILURE__\n\n"

    resp = Response(stream_with_context(generate()), mimetype='text/event-stream')
    resp.headers['Cache-Control'] = 'no-cache'
    resp.headers['X-Accel-Buffering'] = 'no'
    return resp


@app.route('/api/pdf-to-image/convert', methods=['POST'])
def convert_pdfs_to_images():
    """SSE: download each given PDF, render every page to PNG, then stack those pages into
    a single tall PNG per PDF (one image per PDF, not one image per page)."""
    import requests as req
    import fitz  # PyMuPDF

    data = request.get_json(force=True)
    pdfs = data.get('pdfs', [])  # [{'url':..., 'source':...}, ...]
    try:
        dpi = int(data.get('dpi', 150))
    except (TypeError, ValueError):
        dpi = 150
    dpi = max(72, min(dpi, 400))

    def generate():
        global _pdf_image_cache
        _pdf_image_cache = {}

        total       = len(pdfs)
        done        = 0
        total_pages = 0

        yield f"data: Converting {total} PDF(s) at {dpi} DPI...\n\n"

        zoom   = dpi / 72.0
        matrix = fitz.Matrix(zoom, zoom)

        for entry in pdfs:
            pdf_url = (entry.get('url') or '').strip()
            source  = entry.get('source') or ''
            if not pdf_url:
                continue

            basename = pdf_url_to_basename(pdf_url)
            yield f"data: PDF_START: {pdf_url}\n\n"

            try:
                resp = req.get(pdf_url, timeout=30, stream=True)
                if resp.status_code != 200:
                    raise Exception(f"HTTP {resp.status_code}")
                pdf_bytes = resp.content

                doc = fitz.open(stream=pdf_bytes, filetype='pdf')
                if doc.is_encrypted and not doc.authenticate(''):
                    raise Exception('Encrypted / password-protected PDF')

                page_count = doc.page_count
                yield f"data: PDF_INFO: {pdf_url}|{page_count}\n\n"

                page_pngs = []
                for i in range(page_count):
                    pix = doc.load_page(i).get_pixmap(matrix=matrix)
                    page_pngs.append(pix.tobytes('png'))
                    yield f"data: PAGE_DONE: {pdf_url}|{i + 1}|{page_count}\n\n"
                doc.close()

                if page_count > 1:
                    yield f"data: PDF_MERGING: {pdf_url}|{page_count}\n\n"
                combined = stack_pages_into_one_image(page_pngs)

                _pdf_image_cache[pdf_url] = {
                    'source': source, 'image': combined, 'page_count': page_count,
                    'error': None, 'basename': basename,
                }
                total_pages += page_count
                done += 1
                yield f"data: PDF_DONE: {pdf_url}|{page_count}\n\n"

            except Exception as e:
                _pdf_image_cache[pdf_url] = {
                    'source': source, 'image': None, 'page_count': 0,
                    'error': str(e), 'basename': basename,
                }
                done += 1
                yield f"data: PDF_ERROR: {pdf_url}|{e}\n\n"

            yield f"data: PROGRESS: {done}|{total}\n\n"

        yield f"data: \n\n"
        yield f"data: Done. Converted {done} of {total} PDF(s) into one image each ({total_pages} page(s) merged total).\n\n"
        yield "data: __SUCCESS__\n\n"

    resp = Response(stream_with_context(generate()), mimetype='text/event-stream')
    resp.headers['Cache-Control'] = 'no-cache'
    resp.headers['X-Accel-Buffering'] = 'no'
    return resp


@app.route('/api/pdf-to-image/thumbnail')
def pdf_to_image_thumbnail():
    pdf_url = request.args.get('url', '').strip()
    entry = _pdf_image_cache.get(pdf_url)
    if not entry or not entry.get('image'):
        return 'No image available', 404
    return Response(entry['image'], mimetype='image/png')


@app.route('/api/pdf-to-image/download')
def pdf_to_image_download():
    """Download one PDF's single merged image."""
    pdf_url = request.args.get('url', '').strip()
    entry = _pdf_image_cache.get(pdf_url)
    if not entry or not entry.get('image'):
        return 'No image available', 404

    return Response(
        entry['image'],
        mimetype='image/png',
        headers={'Content-Disposition': f'attachment; filename="{entry["basename"]}.png"'}
    )


@app.route('/api/pdf-to-image/export-zip', methods=['POST'])
def pdf_to_image_export_zip():
    """Zip every cached converted image — one image per PDF — (or a given subset of PDF URLs)."""
    import zipfile

    data = request.get_json(silent=True) or {}
    urls = data.get('urls') or list(_pdf_image_cache.keys())

    buf = io.BytesIO()
    used_names = set()
    any_written = False
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for pdf_url in urls:
            entry = _pdf_image_cache.get(pdf_url)
            if not entry or not entry.get('image'):
                continue
            name = f"{entry['basename']}.png"
            final, n = name, 2
            while final in used_names:
                stem, dot, ext = name.rpartition('.')
                final = f"{stem}_{n}{dot}{ext}" if dot else f"{name}_{n}"
                n += 1
            used_names.add(final)
            zf.writestr(final, entry['image'])
            any_written = True

    if not any_written:
        return 'No converted images to export', 400

    buf.seek(0)
    return send_file(buf, mimetype='application/zip', as_attachment=True, download_name='pdf_images.zip')


@app.route('/accessibility-checker')
def accessibility_checker_page():
    return render_template('accessibility_checker.html')


@app.route('/api/accessibility-scan')
def run_accessibility_scan():
    """SSE: crawl a page or a whole site. For every page found, run an axe-core WCAG scan
    and an alt-tag pass, and collect any PDFs for a separate phase-2 check."""
    import requests as req
    from bs4 import BeautifulSoup
    from urllib.parse import urlparse, urljoin
    from collections import deque
    import time, random, json

    url        = request.args.get('url', '').strip()
    mode       = request.args.get('mode', 'page')   # 'page' | 'site'
    find_axe   = request.args.get('axe', 'true').lower() == 'true'
    find_alt   = request.args.get('alt', 'true').lower() == 'true'

    def err(msg):
        yield f"data: {msg}\n\n"
        yield "data: __FAILURE__\n\n"

    if not url or not url.startswith('http'):
        return Response(stream_with_context(err('ERROR: Invalid or missing URL')), mimetype='text/event-stream')

    USER_AGENTS = [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15',
        'Mozilla/5.0 (X11; Linux x86_64; rv:120.0) Gecko/20100101 Firefox/120.0',
    ]

    def generate():
        global _a11y_alt_findings
        _a11y_alt_findings = []

        parsed_start = urlparse(url)
        base_domain  = parsed_start.netloc.lower().removeprefix('www.')

        visited       = set()
        queued        = {url} if mode == 'site' else set()
        queue         = deque([url])
        found_pdfs    = set()
        page_scores   = []
        pages_crawled = 0

        yield f"data: Starting accessibility scan of {url}\n\n"
        yield f"data: Mode: {mode}\n\n"

        try:
            while queue:
                current = queue.popleft()
                if current in visited:
                    continue
                visited.add(current)
                pages_crawled += 1

                yield f"data: PAGE: {pages_crawled}\n\n"
                yield f"data: [{pages_crawled}] {current}\n\n"

                try:
                    headers = {
                        'User-Agent': random.choice(USER_AGENTS),
                        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
                        'Accept-Language': 'en-US,en;q=0.9',
                        'Connection': 'keep-alive',
                        'Upgrade-Insecure-Requests': '1',
                    }
                    resp = req.get(current, headers=headers, timeout=10, allow_redirects=True)
                    if resp.status_code != 200:
                        yield f"data: Skipped (HTTP {resp.status_code}): {current}\n\n"
                        continue

                    content_type = resp.headers.get('Content-Type', '')
                    if 'html' not in content_type:
                        continue

                    soup = BeautifulSoup(resp.text, 'html.parser')
                    title_tag  = soup.find('title')
                    page_title = title_tag.get_text().strip()[:120] if title_tag else ''
                    yield f"data: SCANNED: {current}|{page_title}\n\n"

                    if find_alt:
                        page_alt_findings = extract_alt_tag_findings(soup, current)
                        _a11y_alt_findings.extend(page_alt_findings)
                        for finding in page_alt_findings:
                            yield f"data: IMGALT: {json.dumps(finding, ensure_ascii=True)}\n\n"

                    if find_axe:
                        yield f"data: Running WCAG scan — {current}\n\n"
                        axe_result = run_axe_on_url(current)
                        score = score_from_axe(axe_result)
                        if score is not None:
                            page_scores.append(score)
                        payload = {
                            'url': current, 'score': score,
                            'violations': axe_result.get('violations', []),
                            'error': axe_result.get('error'),
                        }
                        yield f"data: AXE: {json.dumps(payload, ensure_ascii=True)}\n\n"

                    for a in soup.find_all('a', href=True):
                        href = a['href'].strip()
                        if not href or href.startswith('mailto:') or href.startswith('javascript:'):
                            continue
                        full = urljoin(current, href).split('#')[0].rstrip('/')
                        if not full.startswith('http'):
                            continue

                        data_file = (a.get('data-file-name') or '').strip()
                        url_path  = full.lower().split('?')[0]
                        is_pdf    = url_path.endswith('.pdf') or data_file.lower().endswith('.pdf')
                        if is_pdf:
                            if full not in found_pdfs:
                                found_pdfs.add(full)
                                yield f"data: PDF: {full}|{current}\n\n"
                            continue

                        if mode == 'site':
                            link_domain = urlparse(full).netloc.lower().removeprefix('www.')
                            if link_domain == base_domain and full not in visited and full not in queued:
                                queued.add(full)
                                queue.append(full)

                except Exception as e:
                    yield f"data: Error ({current}): {e}\n\n"

                time.sleep(0.1)

            overall = compute_overall_a11y_score(page_scores, [])
            yield f"data: \n\n"
            yield (
                f"data: Done. {pages_crawled} page(s) scanned, {len(_a11y_alt_findings)} image(s) checked, "
                f"{len(found_pdfs)} PDF(s) found. Overall page score: {overall if overall is not None else 'N/A'}.\n\n"
            )
            yield "data: __SUCCESS__\n\n"
        finally:
            close_axe_browser()

    resp = Response(stream_with_context(generate()), mimetype='text/event-stream')
    resp.headers['Cache-Control'] = 'no-cache'
    resp.headers['X-Accel-Buffering'] = 'no'
    return resp


@app.route('/api/accessibility-pdf-check', methods=['POST'])
def run_accessibility_pdf_check():
    """SSE: download each given PDF and run PDF/UA structural accessibility checks —
    run as a separate phase after the page crawl completes, per its own explicit trigger."""
    import requests as req
    import json

    data     = request.get_json(force=True)
    pdfs     = data.get('pdfs', [])  # [{'url':..., 'source':...}, ...]
    use_ai   = bool(data.get('ai'))
    model    = (data.get('model') or '').strip() or None
    ollama_url     = get_ollama_url()
    OLLAMA_TIMEOUT = int(os.getenv('OLLAMA_TIMEOUT', '300'))

    def generate():
        global _a11y_pdf_results
        _a11y_pdf_results = {}

        total = len(pdfs)
        done  = 0
        yield f"data: Checking {total} PDF(s) for accessibility...\n\n"

        for entry in pdfs:
            pdf_url = (entry.get('url') or '').strip()
            source  = entry.get('source') or ''
            if not pdf_url:
                continue

            yield f"data: PDF_START: {pdf_url}\n\n"
            try:
                resp = req.get(pdf_url, timeout=30, stream=True)
                if resp.status_code != 200:
                    raise Exception(f"HTTP {resp.status_code}")
                pdf_bytes = resp.content

                result = check_pdf_accessibility(
                    pdf_bytes, pdf_url, use_ai=use_ai, model=model,
                    ollama_url=ollama_url, timeout=OLLAMA_TIMEOUT,
                )
                result['source'] = source
                _a11y_pdf_results[pdf_url] = result
                done += 1
                yield f"data: PDF_RESULT: {json.dumps(result, ensure_ascii=True)}\n\n"

            except Exception as e:
                result = {'url': pdf_url, 'source': source, 'score': 0, 'error': str(e)}
                _a11y_pdf_results[pdf_url] = result
                done += 1
                yield f"data: PDF_RESULT: {json.dumps(result, ensure_ascii=True)}\n\n"

            yield f"data: PROGRESS: {done}|{total}\n\n"

        pdf_scores = [r.get('score') for r in _a11y_pdf_results.values() if not r.get('error')]
        yield f"data: \n\n"
        yield f"data: Done. Checked {done} of {total} PDF(s).\n\n"
        yield "data: __SUCCESS__\n\n"

    resp = Response(stream_with_context(generate()), mimetype='text/event-stream')
    resp.headers['Cache-Control'] = 'no-cache'
    resp.headers['X-Accel-Buffering'] = 'no'
    return resp


@app.route('/api/export/accessibility-alt')
def export_accessibility_alt():
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(fill_type='solid', fgColor='1F3864')
    ok_font     = Font(color='1E7E34', size=10, bold=True)
    bad_font    = Font(color='B8430A', size=10, bold=True)
    center      = Alignment(horizontal='center', vertical='center')
    left        = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin        = Side(style='thin', color='BFBFBF')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Alt Tags'

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 70

    headers = ['#', 'Page URL', 'Image Src', 'Has Alt', 'Alt Text', 'Source Code Text']
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center; cell.border = border
    ws.row_dimensions[1].height = 20

    for i, finding in enumerate(_a11y_alt_findings, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).alignment = center
        c2 = ws.cell(row=row, column=2, value=finding.get('pageUrl', '')); c2.font = Font(size=10); c2.alignment = left
        c3 = ws.cell(row=row, column=3, value=finding.get('src', '')); c3.font = Font(color='1155CC', size=10); c3.alignment = left
        has_alt = finding.get('hasAlt')
        c4 = ws.cell(row=row, column=4, value='Yes' if has_alt else 'Missing')
        c4.font = ok_font if has_alt else bad_font; c4.alignment = center
        c5 = ws.cell(row=row, column=5, value=finding.get('alt', '')); c5.font = Font(size=10); c5.alignment = left
        c6 = ws.cell(row=row, column=6, value=finding.get('source', '')); c6.font = Font(size=9); c6.alignment = left
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
        ws.row_dimensions[row].height = 16

    ws.freeze_panes = 'A2'
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='accessibility_alt_tags.xlsx'
    )


@app.route('/api/export/accessibility-pdf')
def export_accessibility_pdf():
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(fill_type='solid', fgColor='1F3864')
    ok_font     = Font(color='1E7E34', size=10, bold=True)
    bad_font    = Font(color='B8430A', size=10, bold=True)
    center      = Alignment(horizontal='center', vertical='center')
    left        = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin        = Side(style='thin', color='BFBFBF')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PDF Accessibility'

    widths = [6, 60, 50, 10, 10, 10, 10, 16, 16, 50]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(ord('A') + i)].width = w

    headers = ['#', 'PDF URL', 'Source Page', 'Score', 'Tagged', 'Lang', 'Title',
               'Figures w/ Alt', 'Headings', 'AI Notes']
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center; cell.border = border
    ws.row_dimensions[1].height = 20

    def yn(cell, value):
        cell.value = 'Yes' if value else 'No'
        cell.font = ok_font if value else bad_font
        cell.alignment = center

    for i, r in enumerate(_a11y_pdf_results.values(), 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).alignment = center
        c2 = ws.cell(row=row, column=2, value=r.get('url', '')); c2.font = Font(color='1155CC', size=10); c2.alignment = left
        c3 = ws.cell(row=row, column=3, value=r.get('source', '')); c3.font = Font(size=10); c3.alignment = left
        c4 = ws.cell(row=row, column=4, value=r.get('score', 0)); c4.alignment = center
        c4.font = ok_font if (r.get('score') or 0) >= 80 else (bad_font if (r.get('score') or 0) < 50 else Font(size=10, bold=True))
        yn(ws.cell(row=row, column=5), r.get('tagged'))
        yn(ws.cell(row=row, column=6), r.get('hasLang'))
        yn(ws.cell(row=row, column=7), r.get('hasTitle'))
        figs = f"{r.get('figuresWithAlt', 0)}/{r.get('figureCount', 0)}"
        ws.cell(row=row, column=8, value=figs).alignment = center
        yn(ws.cell(row=row, column=9), r.get('hasHeadings'))
        c10 = ws.cell(row=row, column=10, value=r.get('error') or r.get('aiNotes') or ''); c10.font = Font(size=9); c10.alignment = left
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = border
        ws.row_dimensions[row].height = 16

    ws.freeze_panes = 'A2'
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='accessibility_pdf_report.xlsx'
    )


ACCESSIBILITY_REPORT_LOGO = (
    'https://resources.finalsite.net/images/f_auto,q_auto,t_image_size_1/'
    'v1767801228/ccsdnet/r08p1uzjqujbmeyg2a7g/'
    'WeAreCCSD_DestinationDistrict_OneLine_Vertical_White.png'
)
ACCESSIBILITY_REPORT_HEADER_COLOR = '#2d5b96'
ALT_MISSING_REPORT_LIMIT = 150


@app.route('/api/export/accessibility-report', methods=['POST'])
def export_accessibility_report():
    """Renders a branded PDF summary report (score graphic, PDF-accessibility stat, and
    supporting tables) from whatever the client currently has in memory for this scan,
    by printing an HTML template to PDF with the shared headless browser."""
    from datetime import datetime

    data   = request.get_json(force=True)
    pages  = data.get('pages', [])
    alt    = data.get('alt', {})
    pdfs   = data.get('pdfs', [])
    scores = data.get('scores', {})

    overall_score        = scores.get('overall')
    pdf_inaccessible_pct = scores.get('pdfInaccessiblePct')

    alt_missing_list = alt.get('missingList', [])
    alt_missing_truncated = len(alt_missing_list) > ALT_MISSING_REPORT_LIMIT
    alt_missing_list = alt_missing_list[:ALT_MISSING_REPORT_LIMIT]

    html = render_template(
        'accessibility_report.html',
        url=data.get('url', ''),
        mode=data.get('mode', 'page'),
        generated_at=datetime.now().strftime('%B %d, %Y at %I:%M %p'),
        overall_score=overall_score,
        overall_band=score_band(overall_score),
        pages_scanned=len(pages),
        images_checked=alt.get('total', 0),
        alt_has=alt.get('has', 0),
        alt_missing=alt.get('missing', 0),
        alt_pass_rate=scores.get('altPassRate'),
        pdf_total=scores.get('pdfTotal', 0),
        pdf_inaccessible=scores.get('pdfInaccessible', 0),
        pdf_inaccessible_pct=pdf_inaccessible_pct,
        pdf_band=pct_band_inverse(pdf_inaccessible_pct),
        pages=pages,
        alt_missing_list=alt_missing_list,
        alt_missing_truncated=alt_missing_truncated,
        pdfs=pdfs,
        band_colors=BAND_COLORS,
        score_band=score_band,
        logo_url=ACCESSIBILITY_REPORT_LOGO,
        header_color=ACCESSIBILITY_REPORT_HEADER_COLOR,
    )

    browser = get_axe_browser()
    page = browser.new_page()
    try:
        page.set_content(html, wait_until='networkidle')
        pdf_bytes = page.pdf(
            format='Letter',
            print_background=True,
            margin={'top': '0.3in', 'bottom': '0.6in', 'left': '0.35in', 'right': '0.35in'},
            display_header_footer=True,
            header_template='<span></span>',
            footer_template=(
                '<div style="font-size:8px; width:100%; text-align:center; color:#888;">'
                'Page <span class="pageNumber"></span> of <span class="totalPages"></span></div>'
            ),
        )
    finally:
        page.close()

    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=True,
        download_name='accessibility_report.pdf',
    )


@app.route('/site-architecture')
def site_architecture_page():
    return render_template('site_architecture.html')


def _slugify(text):
    """Fallback slug for pages with no usable path segment (e.g. query-string-routed
    WordPress pages like ?page_id=1697) — derived from the page title instead."""
    slug = re.sub(r"[^a-zA-Z0-9]+", '-', text or '').strip('-').lower()
    return slug[:80]


def _department_label_from_url(url):
    """Derive a human-readable label from a /departments/... URL's final path segment(s)."""
    from urllib.parse import urlparse
    path = urlparse(url).path.strip('/')
    segments = [s for s in path.split('/') if s]
    if segments and segments[0].lower() == 'departments':
        segments = segments[1:]
    if not segments:
        return 'Departments (home)'
    return segments[-1].replace('-', ' ').replace('_', ' ').strip().title()


def _normalize_match_text(text):
    text = text.lower()
    text = re.sub(r'[^a-z0-9\s]', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


# Noise lines that show up on a "Departmental Listing" index page around the
# actual name...page# entries (running headers, A/B/C section letters, footer rev line).
_DIRECTORY_INDEX_NOISE = {'DEPARTMENT', 'PAGE#', 'DEPARTMENTAL LISTING'}
_DIRECTORY_INDEX_ENTRY_RE   = re.compile(r'^(.+?)[.\s]{2,}(\d{1,4})$')
_DIRECTORY_INDEX_LETTER_RE  = re.compile(r'^[A-Z]$')
_DIRECTORY_INDEX_REV_RE     = re.compile(r'^REV\.?\s*\d')


def _extract_directory_departments(pdf_bytes):
    """Pull department names out of a CCSD-style telephone directory PDF.

    These directories include an alphabetical "Departmental Listing" index (name +
    dot-leader + page number) ahead of the actual staff rows — that index is a far
    cleaner source of department names than trying to parse phone numbers back out
    of the tabular position/name/phone/fax rows on the body pages.
    """
    import fitz  # PyMuPDF

    doc = fitz.open(stream=pdf_bytes, filetype='pdf')
    try:
        departments = []
        seen = set()
        for page in doc:
            text = page.get_text('text')
            page_upper = text.upper()
            # Require both markers — "DEPARTMENTAL LISTING" alone also shows up as a
            # cross-reference on the table-of-contents page, which isn't an index page.
            if 'DEPARTMENTAL LISTING' not in page_upper or 'PAGE#' not in page_upper:
                continue  # not an index page — skip (avoids body-page position/phone rows)

            buffer = ''
            for raw_line in text.split('\n'):
                line = raw_line.strip()
                if not line:
                    continue
                upper = line.upper()
                if (upper in _DIRECTORY_INDEX_NOISE
                        or _DIRECTORY_INDEX_LETTER_RE.match(line)
                        or _DIRECTORY_INDEX_REV_RE.match(upper)):
                    buffer = ''
                    continue
                buffer = f'{buffer} {line}'.strip() if buffer else line
                if len(buffer) > 200:
                    buffer = ''  # runaway accumulation without a matching tail — drop and resync
                    continue
                m = _DIRECTORY_INDEX_ENTRY_RE.match(buffer)
                if m:
                    name = re.sub(r'\s+', ' ', m.group(1)).strip(' .')
                    key = name.upper()
                    if name and key not in seen:
                        seen.add(key)
                        departments.append(name)
                    buffer = ''
        return departments
    finally:
        doc.close()


@app.route('/api/site-architecture/extract-directory-pdf', methods=['POST'])
def extract_directory_pdf():
    uploaded = request.files.get('file')
    if not uploaded or not uploaded.filename:
        return jsonify({'ok': False, 'error': 'No file uploaded'}), 400

    try:
        departments = _extract_directory_departments(uploaded.read())
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Could not read PDF: {e}'}), 400

    if not departments:
        return jsonify({
            'ok': False,
            'error': 'No "Departmental Listing" index found in this PDF — it may use a '
                     'different layout than expected.',
        }), 200

    return jsonify({'ok': True, 'departments': departments, 'count': len(departments)})


@app.route('/api/site-architecture/match-departments', methods=['POST'])
def match_departments():
    import difflib
    import json

    data        = request.get_json(force=True) or {}
    departments = [d.strip() for d in data.get('departments', []) if d and d.strip()]

    def generate():
        if not departments:
            yield "data: ERROR: No department names provided\n\n"
            yield "data: __FAILURE__\n\n"
            return

        dept_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'scraper_departments.txt')
        candidates = []
        seen = set()
        try:
            with open(dept_file, 'r', encoding='utf-8') as f:
                for line in f:
                    url = line.strip()
                    if not url:
                        continue
                    dedupe_key = url.split('://', 1)[-1].rstrip('/').lower()
                    if dedupe_key in seen:
                        continue
                    seen.add(dedupe_key)
                    label = _department_label_from_url(url)
                    if label == 'Departments (home)':
                        continue  # the bare /departments index isn't an actual department page
                    candidates.append((url, label))
        except FileNotFoundError:
            yield "data: ERROR: scraper_departments.txt not found — run the main scraper first\n\n"
            yield "data: __FAILURE__\n\n"
            return

        norm_candidates = [(url, label, _normalize_match_text(label)) for url, label in candidates]

        for dept in departments:
            norm_dept = _normalize_match_text(dept)
            best_url, best_label, best_score = None, None, 0.0
            for url, label, norm_label in norm_candidates:
                score = difflib.SequenceMatcher(None, norm_dept, norm_label).ratio()
                if score > best_score:
                    best_score, best_url, best_label = score, url, label
            payload = json.dumps({
                'department': dept,
                'matchedUrl': best_url if best_score >= 0.5 else None,
                'label': best_label if best_score >= 0.5 else None,
                'score': round(best_score, 3),
            })
            yield f"data: MATCH: {payload}\n\n"

        yield "data: __SUCCESS__\n\n"

    resp = Response(stream_with_context(generate()), mimetype='text/event-stream')
    resp.headers['Cache-Control'] = 'no-cache'
    resp.headers['X-Accel-Buffering'] = 'no'
    return resp


@app.route('/api/site-architecture/process', methods=['POST'])
def process_site_architecture():
    import requests as req
    from bs4 import BeautifulSoup
    from urllib.parse import urlparse
    import random, json

    data            = request.get_json(force=True) or {}
    urls            = [u.strip() for u in data.get('urls', []) if u and u.strip()]
    target_domain   = (data.get('targetDomain') or '').strip().rstrip('/')
    requested_model = (data.get('model') or '').strip()
    departments     = [d.strip() for d in data.get('departments', []) if d and d.strip()]
    target_host     = urlparse(target_domain).netloc.lower().removeprefix('www.')
    ollama_url      = get_ollama_url()
    OLLAMA_TIMEOUT  = int(os.getenv('OLLAMA_TIMEOUT', '300'))

    # A large district site's IA, grounded in how Chicago Public Schools (cps.edu) — a
    # real, well-organized big-district site — actually structures its content. Used as
    # a pattern for Ollama to follow rather than inventing an unrelated structure per page.
    SITE_STRUCTURE_GUIDE = (
        "Model the new site's top-level folder structure after how large school district "
        "websites are organized (e.g. Chicago Public Schools' cps.edu). Prefer nesting "
        "pages under these top-level sections rather than inventing unrelated new ones:\n"
        "  /about/ — district info, leadership, board of trustees (e.g. /about/school-board/), "
        "policies, finance, contact, data/stats\n"
        "  /about/departments/ — individual department pages (see the official department "
        "list below, if provided)\n"
        "  /academics/ — curriculum, programs, assessments, course info, enrichment\n"
        "  /schools/ — school locator, enrollment/zoning, individual schools\n"
        "  /families/ — parent/student resources, health, transportation, meals, safety\n"
        "  /calendar/ — calendars and events\n"
        "  /careers/ — employment/jobs\n"
        "  /news/ — press releases, news, media\n"
    )

    DEPARTMENT_VOCAB_HINT = ''
    if departments:
        dept_list = '\n'.join(f'  - {d}' for d in departments[:250])
        DEPARTMENT_VOCAB_HINT = (
            "\nThe following are the OFFICIAL department names for this district (from its "
            "phone directory). If a page's subject matter clearly matches one of these "
            "departments, its folder MUST be /about/departments/<kebab-case of that exact "
            "department name>/ — use the department name as given below, don't invent a "
            "different name for it:\n" + dept_list + "\n"
        )

    USER_AGENTS = [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15',
        'Mozilla/5.0 (X11; Linux x86_64; rv:120.0) Gecko/20100101 Firefox/120.0',
    ]

    def resolve_ollama_model():
        if requested_model:
            return requested_model
        running, installed = list_ollama_models(ollama_url)
        if running:
            return running[0]
        if installed:
            return installed[0]
        return None

    CHUNK_SIZE = 20  # pages per Ollama batch-categorization call
    FOLDER_REGISTRY_LIMIT = 40  # cap prompt growth on very large batches

    def build_registry_hint(folder_registry):
        if not folder_registry:
            return ''
        lines = [
            f"  {folder} — e.g. {'; '.join(examples[:3])}"
            for folder, examples in list(folder_registry.items())[:FOLDER_REGISTRY_LIMIT]
        ]
        return (
            "\n\nFolders already assigned to other pages in this batch — if a page's "
            "subject matter clearly matches one of them, use that EXACT folder path "
            "(character-for-character); only propose a new folder if none of these fit:\n"
            + '\n'.join(lines)
        )

    def get_ollama_batch_folders(model, chunk, registry_hint):
        """Ask Ollama to assign a FOLDER to every page in `chunk` at once — the page's own
        slug/filename is preserved as-is by the caller, never rewritten by the model.
        Returns {normalized_url: folder}."""
        pages_block = '\n'.join(
            f"- URL: {p['url']}\n"
            f"  Slug (fixed — do NOT include in folder): {p['slug']}\n"
            f"  Title: {p['title'] or '(none)'}\n"
            f"  Summary: {p['snippet'] or '(none)'}"
            for p in chunk
        )
        prompt = (
            "You are an information architect for K-12 school district websites. Below is a "
            "batch of pages from an existing site. For EACH page, decide which FOLDER it "
            "should live under on a newly re-organized site — everything in the URL path "
            "EXCEPT the page's own fixed slug (given per-page below), which must NOT be "
            "included in your folder answer. Use lowercase kebab-case folder segments. Pages "
            "covering the same subject matter MUST get the exact same folder — compare pages "
            "against each other (and against the already-assigned folders, if any) before "
            "deciding.\n\n" + SITE_STRUCTURE_GUIDE + DEPARTMENT_VOCAB_HINT
            + registry_hint +
            "\n\nRespond with ONLY a JSON array, one object per page listed below, in exactly "
            "this form and nothing else (no markdown, no commentary):\n"
            '[{"url": "<page url copied exactly as given>", "folder": "<folder path, '
            'starting and ending with />"}]\n\n'
            f"Pages:\n{pages_block}"
        )
        r = req.post(
            f"{ollama_url}/api/generate",
            json={'model': model, 'prompt': prompt, 'stream': False},
            timeout=OLLAMA_TIMEOUT,
        )
        r.raise_for_status()
        text = r.json().get('response', '').strip()
        text = re.sub(r'^```(?:json)?\s*', '', text)
        text = re.sub(r'\s*```$', '', text).strip()
        m = re.search(r'\[[\s\S]*\]', text)
        raw = m.group() if m else text
        raw = re.sub(r',\s*([}\]])', r'\1', raw)  # trailing commas some models emit
        items = json.loads(raw)

        assignments = {}
        for item in items:
            item_url = (item.get('url') or '').strip()
            folder   = (item.get('folder') or '').strip()
            if not item_url or not folder:
                continue
            if not folder.startswith('/'):
                folder = '/' + folder
            if not folder.endswith('/'):
                folder += '/'
            assignments[item_url.rstrip('/')] = folder
        return assignments

    def generate():
        yield f"data: Received {len(urls)} link(s). Starting...\n\n"

        if not urls:
            yield "data: ERROR: No URLs provided\n\n"
            yield "data: __FAILURE__\n\n"
            return
        if not target_domain:
            yield "data: ERROR: No target domain provided\n\n"
            yield "data: __FAILURE__\n\n"
            return

        try:
            resolved_model = resolve_ollama_model()
        except Exception as e:
            yield f"data: ERROR: Could not reach Ollama at {ollama_url} — {e}\n\n"
            yield "data: __FAILURE__\n\n"
            return

        if not resolved_model:
            yield f"data: ERROR: No Ollama model found at {ollama_url}\n\n"
            yield "data: __FAILURE__\n\n"
            return

        yield f"data: Using Ollama model: {resolved_model}\n\n"

        # ── Phase 1: scrape every page for its title + a short content snippet ──────
        yield f"data: Phase 1/2 — scraping {len(urls)} page(s)...\n\n"
        pages_to_categorize = []

        for i, url in enumerate(urls, 1):
            yield f"data: Scraping ({i}/{len(urls)}): {url}\n\n"
            try:
                headers = {
                    'User-Agent': random.choice(USER_AGENTS),
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                    'Accept-Language': 'en-US,en;q=0.9',
                    'Connection': 'keep-alive',
                }
                resp = req.get(url, headers=headers, timeout=15, allow_redirects=True)
                if resp.status_code != 200:
                    payload = json.dumps({'originalUrl': url, 'suggestedUrl': None, 'title': None,
                                           'error': f'HTTP {resp.status_code}'})
                    yield f"data: RESULT: {payload}\n\n"
                    continue

                content_type = resp.headers.get('Content-Type', '')
                if 'html' not in content_type:
                    payload = json.dumps({'originalUrl': url, 'suggestedUrl': None, 'title': None,
                                           'error': f'Not HTML ({content_type})'})
                    yield f"data: RESULT: {payload}\n\n"
                    continue

                raw_html   = resp.text
                soup       = BeautifulSoup(raw_html, 'html.parser')
                title_tag  = soup.find('title')
                page_title = title_tag.get_text().strip()[:150] if title_tag else ''

                content_html = extract_content_html(raw_html, url)
                content_text = BeautifulSoup(content_html, 'html.parser').get_text(separator=' ', strip=True)
                content_text = re.sub(r'\s+', ' ', content_text).strip()

                if not content_text:
                    payload = json.dumps({'originalUrl': url, 'suggestedUrl': None, 'title': page_title,
                                           'error': 'No extractable text content'})
                    yield f"data: RESULT: {payload}\n\n"
                    continue

                parsed_url    = urlparse(url)
                path_segments = [s for s in parsed_url.path.split('/') if s]
                page_slug     = path_segments[-1] if path_segments else ''
                page_host     = parsed_url.netloc.lower().removeprefix('www.')

                if not page_slug and not parsed_url.query and page_host == target_host:
                    # True homepage of the site being reorganized (its own root, no path,
                    # no query) — nothing to categorize. A *different* ccsd.net subdomain's
                    # own root (e.g. aarsi.ccsd.net) is NOT this — it's just another page
                    # that needs categorizing like anything else.
                    payload = json.dumps({'originalUrl': url, 'suggestedUrl': target_domain + '/',
                                           'title': page_title, 'error': None})
                    yield f"data: RESULT: {payload}\n\n"
                    continue

                if not page_slug:
                    # No path segment — a bare subdomain root, or a query-string-routed page
                    # (e.g. WordPress ?page_id=1697). There's no real slug to preserve, so
                    # derive one from the title (or host, as a last resort) instead.
                    page_slug = _slugify(page_title) or _slugify(parsed_url.query) or _slugify(page_host)

                pages_to_categorize.append({
                    'url': url, 'slug': page_slug, 'title': page_title,
                    'snippet': content_text[:300],
                })

            except Exception as e:
                payload = json.dumps({'originalUrl': url, 'suggestedUrl': None, 'title': None, 'error': str(e)})
                yield f"data: RESULT: {payload}\n\n"

        if not pages_to_categorize:
            yield "data: __SUCCESS__\n\n"
            return

        # ── Phase 2: batch-categorize with Ollama, in chunks, sharing a folder registry ──
        chunks = [pages_to_categorize[i:i + CHUNK_SIZE] for i in range(0, len(pages_to_categorize), CHUNK_SIZE)]
        yield (f"data: Phase 2/2 — categorizing {len(pages_to_categorize)} page(s) in "
               f"{len(chunks)} batch(es) using {resolved_model}...\n\n")

        folder_registry = {}  # folder -> [example titles], grows across chunks

        def categorize_with_retry(chunk):
            """Try to categorize `chunk` in one Ollama call. If the model's response can't
            be parsed (bad/truncated JSON is common on larger batches with smaller local
            models), split the chunk in half and retry each half — down to individual pages
            if needed — instead of letting one bad batch wipe out every page in it.
            Returns ({normalized_url: folder}, {normalized_url: error_message})."""
            try:
                return get_ollama_batch_folders(resolved_model, chunk, build_registry_hint(folder_registry)), {}
            except Exception as e:
                if len(chunk) <= 1:
                    return {}, {p['url'].rstrip('/'): str(e) for p in chunk}
                mid = len(chunk) // 2
                a1, e1 = categorize_with_retry(chunk[:mid])
                a2, e2 = categorize_with_retry(chunk[mid:])
                return {**a1, **a2}, {**e1, **e2}

        for ci, chunk in enumerate(chunks, 1):
            yield f"data: Categorizing batch {ci}/{len(chunks)} ({len(chunk)} page(s))...\n\n"
            assignments, batch_errors = categorize_with_retry(chunk)

            for p in chunk:
                key = p['url'].rstrip('/')
                folder = assignments.get(key)
                if not folder:
                    err = batch_errors.get(key, 'Ollama did not return a folder for this page')
                    payload = json.dumps({'originalUrl': p['url'], 'suggestedUrl': None, 'title': p['title'],
                                           'error': err})
                    yield f"data: RESULT: {payload}\n\n"
                    continue

                # Defensive: if the model echoed the slug back onto the end of the folder
                # anyway, strip that trailing copy so it isn't duplicated below.
                slug_suffix = f"{p['slug']}/"
                if folder.endswith('/' + slug_suffix):
                    folder = folder[: -len(slug_suffix)]

                # The page's own slug is preserved exactly as it was — only the folder
                # path leading up to it is reorganized.
                suggested_url = target_domain + folder + p['slug']
                folder_registry.setdefault(folder, []).append(p['title'] or p['slug'])
                payload = json.dumps({'originalUrl': p['url'], 'suggestedUrl': suggested_url,
                                       'title': p['title'], 'error': None})
                yield f"data: RESULT: {payload}\n\n"

        yield "data: __SUCCESS__\n\n"

    resp = Response(stream_with_context(generate()), mimetype='text/event-stream')
    resp.headers['Cache-Control'] = 'no-cache'
    resp.headers['X-Accel-Buffering'] = 'no'
    return resp


if __name__ == '__main__':
    app.run(debug=True, port=5002)